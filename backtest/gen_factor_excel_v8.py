"""M30 RSI v8 — 5因子阶梯分 + ls-ss净得分 完整分析表"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "全组合"

hdr_font = Font(bold=True, size=11, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="2F5496")
green = PatternFill("solid", fgColor="C6EFCE")
red = PatternFill("solid", fgColor="FFC7CE")
yellow = PatternFill("solid", fgColor="FFEB9C")
thin = Side(style='thin')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── 因子取值 ──
bb_pos = {"下轨": (1,0), "下轨~中轨": (0,0), "中轨~上轨": (0,0), "上轨": (0,1)}
rsi_val_map = {"<30": (2,0), "30-50": (1,0), "50-70": (0,1), ">70": (0,2)}
rsi_cross_map = {"6>13>27": (2,0), "6>13": (1,0), "6<13": (0,1), "6<13<27": (0,2)}
# MA14: close<MA14 → 回归做多+1; close>MA14 → 回归做空+1
ma_map = {"<MA14(做多回归)": (1,0), ">MA14(做空回归)": (0,1)}
di_map = {"+DI>2×-DI": (2,0), "+DI>-DI": (1,0), "-DI>+DI": (0,1), "-DI>2×+DI": (0,2)}

def is_possible(bb, rsi, cross, ma, di):
    """排除理论上不成立的组合"""
    # BB下轨时RSI不可能>50
    if bb == "下轨" and rsi in ("50-70", ">70"):
        return False
    # BB上轨时RSI不可能<50
    if bb == "上轨" and rsi in ("<30", "30-50"):
        return False
    # BB下轨~中轨时RSI>70极罕见
    if bb == "下轨~中轨" and rsi == ">70":
        return False
    # BB中轨~上轨时RSI<30极罕见
    if bb == "中轨~上轨" and rsi == "<30":
        return False
    # 强金叉(6>13>27)时RSI不可能<50
    if cross == "6>13>27" and rsi in ("<30", "30-50"):
        return False
    # 强死叉(6<13<27)时RSI不可能>50
    if cross == "6<13<27" and rsi in ("50-70", ">70"):
        return False
    # 弱金叉(6>13)时RSI<30罕见
    if cross == "6>13" and rsi == "<30":
        return False
    # 弱死叉(6<13)时RSI>70罕见
    if cross == "6<13" and rsi == ">70":
        return False
    return True

HEADERS = [
    "#", "BB", "RSI值", "RSI交叉", "MA14", "DI",
    "ls分", "ss分", "净分(ls-ss)", "信号",
    "ls明细", "ss明细"
]

for col, h in enumerate(HEADERS, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = Alignment(horizontal='center', wrap_text=True)
    c.border = border

rows = []
seq = 0

for bb, (bb_l, bb_s) in bb_pos.items():
    for rsi, (r_l, r_s) in rsi_val_map.items():
        for cross, (c_l, c_s) in rsi_cross_map.items():
            for ma, (m_l, m_s) in ma_map.items():
                for di, (d_l, d_s) in di_map.items():
                    if not is_possible(bb, rsi, cross, ma, di):
                        continue
                    seq += 1
                    ls = m_l + bb_l + r_l + c_l + d_l
                    ss = m_s + bb_s + r_s + c_s + d_s
                    net = ls - ss

                    if net >= 3:
                        sig = "做多 ✅"
                    elif net <= -3:
                        sig = "做空 ✅"
                    else:
                        sig = "观望 ❌"

                    # 明细
                    ls_detail = []
                    if m_l: ls_detail.append("MA14")
                    if bb_l: ls_detail.append("BB下轨")
                    if r_l: ls_detail.append(f"RSI{r_l}分")
                    if c_l: ls_detail.append(f"金叉{c_l}分")
                    if d_l: ls_detail.append(f"DI{d_l}分")
                    ss_detail = []
                    if m_s: ss_detail.append("MA14")
                    if bb_s: ss_detail.append("BB上轨")
                    if r_s: ss_detail.append(f"RSI{r_s}分")
                    if c_s: ss_detail.append(f"死叉{c_s}分")
                    if d_s: ss_detail.append(f"DI{d_s}分")

                    rows.append([
                        seq, bb, rsi, cross, ma, di,
                        ls, ss, net, sig,
                        "+".join(ls_detail) if ls_detail else "-",
                        "+".join(ss_detail) if ss_detail else "-"
                    ])

for i, r in enumerate(rows, 2):
    for col, val in enumerate(r, 1):
        c = ws.cell(row=i, column=col, value=val)
        c.border = border
        c.alignment = Alignment(horizontal='center', wrap_text=True)
        h = HEADERS[col-1]
        # 净得分绿色(≥3) / 红色(≤-3)
        if h == "净分(ls-ss)" and isinstance(val, int):
            if val >= 3: c.fill = green
            elif val <= -3: c.fill = red
            if val > 0 and val < 3: c.fill = yellow
        if h == "信号":
            if "做多" in str(val): c.fill = green
            elif "做空" in str(val): c.fill = red
            else: c.fill = yellow

# 列宽
for i, w in enumerate([4, 12, 8, 12, 16, 14, 6, 6, 12, 12, 28, 28], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A2"

# ── Sheet2: 评分规则 ──
ws2 = wb.create_sheet("评分规则")
rules = [
    ("因子", "条件", "ls", "ss"),
    ("① MA14回归", "close < MA14", "1", "0"),
    ("", "close > MA14", "0", "1"),
    ("② BB通道", "下轨", "1", "0"),
    ("", "上轨", "0", "1"),
    ("③ RSI值", "< 30", "2", "0"),
    ("", "30 ~ 50", "1", "0"),
    ("", "50 ~ 70", "0", "1"),
    ("", "> 70", "0", "2"),
    ("④ RSI交叉(6/13/27)", "6 > 13 > 27(强金叉)", "2", "0"),
    ("", "6 > 13(弱金叉)", "1", "0"),
    ("", "6 < 13(弱死叉)", "0", "1"),
    ("", "6 < 13 < 27(强死叉)", "0", "2"),
    ("⑤ DI强弱", "+DI > 2×-DI(强多)", "2", "0"),
    ("", "+DI > -DI(弱多)", "1", "0"),
    ("", "-DI > +DI(弱空)", "0", "1"),
    ("", "-DI > 2×+DI(强空)", "0", "2"),
    ("", "", "", ""),
    ("开仓规则", "ls - ss ≥ +3 → 做多", "", ""),
    ("", "ls - ss ≤ -3 → 做空", "", ""),
]
for i, r in enumerate(rules, 1):
    for j, v in enumerate(r, 1):
        c = ws2.cell(row=i, column=j, value=v)
        c.border = border
        c.alignment = Alignment(horizontal='center')
        if i == 1: c.font = hdr_font; c.fill = hdr_fill

ws2.column_dimensions['A'].width = 18
ws2.column_dimensions['B'].width = 26

# ── Sheet3: 可进场场景 ──
ws3 = wb.create_sheet("可进场")
hd3 = ["BB", "RSI", "RSI交叉", "MA14", "DI", "净分", "方向", "关键驱动"]
for col, h in enumerate(hd3, 1):
    c = ws3.cell(row=1, column=col, value=h)
    c.font = hdr_font; c.fill = hdr_fill
    c.alignment = Alignment(horizontal='center', wrap_text=True)
    c.border = border

tradeable = [r for r in rows if r[9] in ("做多 ✅", "做空 ✅")]
for i, r in enumerate(tradeable, 2):
    _, bb, rsi, cross, ma, di, ls, ss, net, sig = r[:10]
    # 关键驱动因子: 找到贡献最大的
    parts = []
    if "强金叉" in cross or "强多" in di: parts.append("趋势强")
    if "下轨" in bb and ("<30" in rsi or "30-50" in rsi): parts.append("超卖")
    if "上轨" in bb and (">70" in rsi or "50-70" in rsi): parts.append("超买")
    if "回归" in ma: parts.append("回归")
    driver = "+".join(parts) if parts else "综合"

    ws3.cell(row=i, column=1, value=bb).border = border
    ws3.cell(row=i, column=2, value=rsi).border = border
    ws3.cell(row=i, column=3, value=cross).border = border
    ws3.cell(row=i, column=4, value=ma).border = border
    ws3.cell(row=i, column=5, value=di).border = border
    ws3.cell(row=i, column=6, value=net).border = border
    ws3.cell(row=i, column=7, value=sig).border = border
    ws3.cell(row=i, column=8, value=driver).border = border

    fill = green if "做多" in sig else red if "做空" in sig else yellow
    ws3.cell(row=i, column=7).fill = fill
    if isinstance(net, int) and abs(net) >= 5:
        ws3.cell(row=i, column=6).fill = green
    for col in range(1, 9):
        ws3.cell(row=i, column=col).alignment = Alignment(horizontal='center')

for i, w in enumerate([12, 8, 12, 16, 14, 6, 12, 20], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.freeze_panes = "A2"

# ── Sheet4: 净分分布统计 ──
ws4 = wb.create_sheet("统计")
ws4.cell(row=1, column=1, value="净分").font = hdr_font; ws4.cell(row=1, column=1).fill = hdr_fill
ws4.cell(row=1, column=2, value="组合数").font = hdr_font; ws4.cell(row=1, column=2).fill = hdr_fill
ws4.cell(row=1, column=3, value="占比").font = hdr_font; ws4.cell(row=1, column=3).fill = hdr_fill

from collections import Counter
net_counts = Counter(r[8] for r in rows)
total = len(rows)
row_i = 2
for net_val in sorted(net_counts.keys()):
    cnt = net_counts[net_val]
    ws4.cell(row=row_i, column=1, value=net_val).border = border
    ws4.cell(row=row_i, column=2, value=cnt).border = border
    ws4.cell(row=row_i, column=3, value=f"{cnt/total*100:.1f}%").border = border
    if net_val >= 3:
        ws4.cell(row=row_i, column=1).fill = green
    elif net_val <= -3:
        ws4.cell(row=row_i, column=1).fill = red
    row_i += 1

ws4.cell(row=row_i, column=1, value="合计").font = hdr_font
ws4.cell(row=row_i, column=2, value=total).font = hdr_font
ws4.cell(row=row_i, column=3, value="100%").font = hdr_font

buy_count = sum(1 for r in rows if r[8] >= 3)
sell_count = sum(1 for r in rows if r[8] <= -3)
row_i += 2
ws4.cell(row=row_i, column=1, value="做多场景(净分≥+3)").font = hdr_font
ws4.cell(row=row_i, column=1).fill = green
ws4.cell(row=row_i, column=2, value=buy_count)
row_i += 1
ws4.cell(row=row_i, column=1, value="做空场景(净分≤-3)").font = hdr_font
ws4.cell(row=row_i, column=1).fill = red
ws4.cell(row=row_i, column=2, value=sell_count)
row_i += 1
ws4.cell(row=row_i, column=1, value="观望场景").font = hdr_font
ws4.cell(row=row_i, column=1).fill = yellow
ws4.cell(row=row_i, column=2, value=total - buy_count - sell_count)

for i, w in enumerate([22, 10, 10], 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

wb.save("D:/backup/BaoBao/PythonProgram/xauusd/backtest/factor_table_v8.xlsx")
print(f"Done! {len(rows)} realistic combos")
print(f"  做多(net≥+3): {buy_count}")
print(f"  做空(net≤-3): {sell_count}")
print(f"  观望: {total - buy_count - sell_count}")
