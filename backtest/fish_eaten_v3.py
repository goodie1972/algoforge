"""
fish_eaten v3 入场改造 · 回测对比
=================================
对比「v2 旧 6 道筛子」与「v3 衰竭打分」在同一份数据上的表现，
并网格扫描 v3 参数（score_min / adx_max / require_pierce / div_lookback）。

入场打分逻辑直接复用 strategies/fish_eaten_entry.py —— 与实盘同一份代码，
保证「回测看到的就是实盘跑的」。

用法：
    python backtest/fish_eaten_v3.py                      # 默认 CSV M30
    python backtest/fish_eaten_v3.py --source db          # 用 indicator_snapshots（数据更新）
    python backtest/fish_eaten_v3.py --tf M15 --sl atr    # 加 1.5×ATR 硬止损（对齐实盘）
    python backtest/fish_eaten_v3.py --min-trades 15      # 过滤样本过少的组合

核心关注指标（不只看 PnL，样本少时 PnL 会被运气单主导）：
    trades   交易笔数
    winrate  胜率
    pnl      净盈亏（0.01 手口径，1 美元/盎司 = 1 美元）
    maxdd    最大回撤
    adv6     入场后 6 根内最大逆向幅度（越大 = 越像在趋势中段挨打）
    cont12   入场后 12 根内价格沿原趋势方向继续推进的幅度（越大 = 入场越早）
"""
import argparse
import json
import os
import sys

import numpy as np
import pandas as pd

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)

from strategies.fish_eaten_entry import EntryParams, score_entry  # noqa: E402

try:
    import talib
except ImportError:
    sys.exit("需要 TA-Lib：请用带 talib 的解释器运行（如 C:\\Python314\\python.exe）")

LOT_OZ = 1.0  # 0.01 手 = 1 盎司 → 价格差 1 美元 = 盈亏 1 美元


# ───────────────────────── 数据加载 ─────────────────────────

def load_csv(tf):
    path = os.path.join(ROOT, 'data', f'XAUUSD_{tf}.csv')
    if not os.path.exists(path):
        sys.exit(f"找不到 {path}")
    df = pd.read_csv(path, names=['time', 'open', 'high', 'low', 'close', 'volume'],
                     skiprows=1, parse_dates=['time'])
    return df


def load_db(tf):
    """从 ohlcv + indicator_snapshots 取数（指标是 DataFactory 已算好的，与实盘同源）"""
    import sqlite3
    db = os.path.join(ROOT, 'data', 'market_data.db')
    if not os.path.exists(db):
        sys.exit(f"找不到 {db}")
    conn = sqlite3.connect(db)
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        "SELECT o.timestamp, o.open, o.high, o.low, o.close, o.volume, i.indicators "
        "FROM ohlcv o LEFT JOIN indicator_snapshots i "
        "ON i.timeframe = o.timeframe AND i.timestamp = o.timestamp "
        "WHERE o.timeframe = ? ORDER BY o.timestamp", (tf,)
    ).fetchall()
    conn.close()
    if not rows:
        sys.exit(f"DB 中无 {tf} 数据")
    recs = []
    for r in rows:
        ind = {}
        if r['indicators']:
            try:
                ind = json.loads(r['indicators'])
            except Exception:
                ind = {}
        recs.append({'time': pd.to_datetime(r['timestamp'], unit='s'),
                     'open': r['open'], 'high': r['high'], 'low': r['low'],
                     'close': r['close'], 'volume': r['volume'], '_ind': ind})
    return pd.DataFrame(recs)


def calc_indicators(df, from_db=False):
    o = df['open'].values.astype(float)
    h = df['high'].values.astype(float)
    l = df['low'].values.astype(float)
    c = df['close'].values.astype(float)
    v = df['volume'].values.astype(float)

    if from_db:
        inds = df['_ind'].tolist()
        def col(name, default=np.nan):
            out = np.full(len(df), np.nan)
            for i, d in enumerate(inds):
                x = d.get(name) if isinstance(d, dict) else None
                if x is not None:
                    try:
                        out[i] = float(x)
                    except (TypeError, ValueError):
                        pass
            return out
        bb = [ (d.get('bb') if isinstance(d, dict) else None) for d in inds ]
        bb_top = np.array([float(b['upper']) if isinstance(b, dict) and 'upper' in b else np.nan for b in bb])
        bb_mid = np.array([float(b['mid']) if isinstance(b, dict) and 'mid' in b else np.nan for b in bb])
        bb_bot = np.array([float(b['lower']) if isinstance(b, dict) and 'lower' in b else np.nan for b in bb])
        df['rsi'] = col('rsi'); df['mfi'] = col('mfi')
        df['adx'] = col('adx'); df['pdi'] = col('pdi'); df['ndi'] = col('ndi')
        df['bb_width'] = col('bb_width')
        df['atr'] = col('atr')
    else:
        df['rsi'] = talib.RSI(c, timeperiod=14)
        df['mfi'] = talib.MFI(h, l, c, v, timeperiod=14)
        df['adx'] = talib.ADX(h, l, c, timeperiod=14)
        df['pdi'] = talib.PLUS_DI(h, l, c, timeperiod=14)
        df['ndi'] = talib.MINUS_DI(h, l, c, timeperiod=14)
        bb_top, bb_mid, bb_bot = talib.BBANDS(c, timeperiod=20, nbdevup=2, nbdevdn=2)
        df['bb_width'] = bb_top - bb_bot
        df['atr'] = talib.ATR(h, l, c, timeperiod=14)

    df['bb_top'] = bb_top
    df['bb_mid'] = bb_mid
    df['bb_bot'] = bb_bot
    df['bb_mid_dir'] = np.where(df['bb_mid'].diff() > 0, 'up',
                                np.where(df['bb_mid'].diff() < 0, 'down', 'flat'))
    return df


# ───────────────────────── 回测引擎 ─────────────────────────

def _maxdd(pnls):
    eq, peak, dd = 0.0, 0.0, 0.0
    for p in pnls:
        eq += p
        peak = max(peak, eq)
        dd = min(dd, eq - peak)
    return abs(dd)


def _diagnostics(bars, i_entry, i_exit, side):
    """入场位置诊断：衡量「是不是在趋势中段追进去的」。

    adv6  = 入场后 6 根内最大逆向幅度（被套深度，越小越好）
    adv12 = 入场后 12 根内最大逆向幅度（趋势延续深度，越小越好 → 越小说明越接近末端）
    mfe12 = 入场后 12 根内最大有利偏移（浮盈潜力，越大越好）
    """
    long_side = (side == 'LONG')
    lo = max(0, i_entry)
    hi = min(len(bars['low']), i_exit + 1)
    if hi <= lo:
        return 0.0, 0.0, 0.0
    lows, highs = bars['low'][lo:hi], bars['high'][lo:hi]
    entry = bars['open'][i_entry]
    if long_side:
        adv6 = entry - lows[:6].min()
        adv12 = entry - lows[:12].min()
        mfe12 = highs[:12].max() - entry
        mae = entry - lows.min()
        mfe = highs.max() - entry
    else:
        adv6 = highs[:6].max() - entry
        adv12 = highs[:12].max() - entry
        mfe12 = entry - lows[:12].min()
        mae = highs.max() - entry
        mfe = entry - lows.min()
    return (round(float(adv6), 2), round(float(adv12), 2), round(float(mfe12), 2),
            round(float(mae), 2), round(float(mfe), 2))


def run_backtest(bars, mode='v3', p: EntryParams = None,
                 time_stop=48, bb_exit=8, sl_mode='none',
                 sl_atr_mult=1.5, sl_min=15.0, struct_exit=0,
                 rsi_os=30, rsi_ob=70, mfi_os=25, mfi_ob=75, bb_entry_offset=5):
    """mode: 'v2' 旧 6 道筛子 / 'v3' 衰竭打分"""
    n = len(bars['close'])
    trades = []
    in_pos = False
    direction = None
    entry_i = entry_price = 0
    rsi_ext = mfi_ext = False
    first_ext_i = -1
    pending = None
    sl_price = None

    for i in range(n):
        rsi, mfi = bars['rsi'][i], bars['mfi'][i]
        adx, pdi, ndi = bars['adx'][i], bars['pdi'][i], bars['ndi'][i]
        bb_top, bb_bot = bars['bb_top'][i], bars['bb_bot'][i]
        op, cl = bars['open'][i], bars['close'][i]

        if np.isnan(rsi) or np.isnan(mfi) or np.isnan(adx) or np.isnan(bb_bot):
            continue

        # ── 执行挂单（下一根开盘成交）──
        if pending and not in_pos:
            entry_price = op
            entry_i = i
            direction = 'LONG' if pending == 'buy' else 'SHORT'
            in_pos = True
            rsi_ext = mfi_ext = False
            first_ext_i = -1
            atr = bars['atr'][i]
            if sl_mode == 'atr' and not np.isnan(atr):
                dist = max(float(atr) * sl_atr_mult, sl_min)
                sl_price = entry_price - dist if direction == 'LONG' else entry_price + dist
            else:
                sl_price = None
            pending = None
            continue

        # ── 持仓中 ──
        if in_pos:
            long_side = (direction == 'LONG')
            if long_side:
                ext_hit_r = rsi >= rsi_ob
                ext_hit_m = mfi >= mfi_ob
                left = (rsi < rsi_ob or mfi < mfi_ob)
                close_cond = cl < bb_top - bb_exit
            else:
                ext_hit_r = rsi <= rsi_os
                ext_hit_m = mfi <= mfi_os
                left = (rsi > rsi_os or mfi > mfi_os)
                close_cond = cl > bb_bot + bb_exit

            if not rsi_ext and ext_hit_r:
                rsi_ext = True
                if first_ext_i == -1:
                    first_ext_i = i
            if not mfi_ext and ext_hit_m:
                mfi_ext = True
                if first_ext_i == -1:
                    first_ext_i = i

            exit_reason = None
            exit_price = None

            # 硬止损（用当根 low/high 触发）
            if sl_price is not None:
                if long_side and bars['low'][i] <= sl_price:
                    exit_reason, exit_price = 'sl', sl_price
                elif (not long_side) and bars['high'][i] >= sl_price:
                    exit_reason, exit_price = 'sl', sl_price

            if exit_reason is None and struct_exit and (i - entry_i) >= 1:
                lo_i = max(0, i - struct_exit)
                win = bars['close'][lo_i:i]
                if win.size:
                    if long_side and cl < np.nanmin(win):
                        exit_reason, exit_price = 'struct_exit', cl   # 创新低：趋势延续，我错了
                    elif (not long_side) and cl > np.nanmax(win):
                        exit_reason, exit_price = 'struct_exit', cl   # 创新高：趋势延续，我错了

            if exit_reason is None:
                both = rsi_ext and mfi_ext
                if not both and first_ext_i != -1 and (i - first_ext_i) >= time_stop:
                    exit_reason, exit_price = 'time_stop', cl
                elif both and left and close_cond:
                    exit_reason, exit_price = 'fish_exit', cl

            if exit_reason:
                pnl = (exit_price - entry_price) if long_side else (entry_price - exit_price)
                adv6, adv12, mfe12, mae, mfe = _diagnostics(bars, entry_i, i, direction)
                trades.append({'i_entry': entry_i, 'i_exit': i,
                               'entry_time': bars['time'][entry_i], 'exit_time': bars['time'][i],
                               'direction': direction, 'entry': round(entry_price, 2),
                               'exit': round(exit_price, 2), 'bars': i - entry_i,
                               'pnl': round(pnl * LOT_OZ, 2), 'reason': exit_reason,
                               'adv6': adv6, 'adv12': adv12, 'mfe12': mfe12,
                               'mae': mae, 'mfe': mfe})
                in_pos = False
                direction = None
                sl_price = None
            continue

        # ── 空仓：入场判定 ──
        if mode == 'v2':
            if adx <= 22 or abs(pdi - ndi) <= 5:
                continue
            if ndi > pdi:
                if (rsi < rsi_os and mfi < mfi_os and cl <= bb_bot + bb_entry_offset
                        and bars['bb_mid_dir'][i] == 'down'):
                    pending = 'buy'
            else:
                if (rsi > rsi_ob and mfi > mfi_ob and cl >= bb_top - bb_entry_offset
                        and bars['bb_mid_dir'][i] == 'up'):
                    pending = 'sell'
        else:
            side = 'LONG' if ndi > pdi else 'SHORT'
            lb = p.div_lookback + 2
            lo = max(0, i - lb + 1)
            ctx = {
                'rsi': rsi, 'rsi_prev': bars['rsi'][i - 1] if i >= 1 else None,
                'mfi': mfi, 'mfi_prev': bars['mfi'][i - 1] if i >= 1 else None,
                'adx': adx, 'adx_prev': bars['adx'][i - 1] if i >= 1 else None,
                'pdi': pdi, 'pdi_prev': bars['pdi'][i - 1] if i >= 1 else None,
                'ndi': ndi, 'ndi_prev': bars['ndi'][i - 1] if i >= 1 else None,
                'bb_lower': bb_bot, 'bb_upper': bb_top,
                'bb_width': bars['bb_width'][i],
                'bb_width_prev': bars['bb_width'][i - 1] if i >= 1 else None,
                'close': cl, 'low': bars['low'][i], 'high': bars['high'][i],
                'closes': [float(x) for x in bars['close'][lo:i + 1] if not np.isnan(x)],
                'rsis': [float(x) for x in bars['rsi'][lo:i + 1] if not np.isnan(x)],
            }
            res = score_entry(side, ctx, p)
            if res['pass']:
                pending = 'buy' if side == 'LONG' else 'sell'
    return trades


def summarize(name, trades):
    if not trades:
        return {'name': name, 'trades': 0, 'wins': 0, 'winrate': 0, 'pnl': 0,
                'maxdd': 0, 'avg_bars': 0, 'adv6': 0, 'adv12': 0, 'mfe12': 0, 'mae': 0, 'mfe': 0, 'avg_pnl': 0}
    pnls = [t['pnl'] for t in trades]
    wins = sum(1 for p in pnls if p > 0)
    return {
        'name': name, 'trades': len(trades), 'wins': wins,
        'winrate': round(wins / len(trades) * 100, 1),
        'pnl': round(sum(pnls), 2),
        'avg_pnl': round(sum(pnls) / len(trades), 2),
        'maxdd': round(_maxdd(pnls), 2),
        'avg_bars': round(sum(t['bars'] for t in trades) / len(trades), 1),
        'adv6': round(sum(t['adv6'] for t in trades) / len(trades), 2),
        'adv12': round(sum(t['adv12'] for t in trades) / len(trades), 2),
        'mfe12': round(sum(t['mfe12'] for t in trades) / len(trades), 2),
        'mae': round(sum(t['mae'] for t in trades) / len(trades), 2),
        'mfe': round(sum(t['mfe'] for t in trades) / len(trades), 2),
        'fish_exit': sum(1 for t in trades if t['reason'] == 'fish_exit'),
        'time_stop': sum(1 for t in trades if t['reason'] == 'time_stop'),
        'sl': sum(1 for t in trades if t['reason'] == 'sl'),
    }


# ───────────────────────── 主流程 ─────────────────────────

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--tf', default='M30')
    ap.add_argument('--source', default='csv', choices=['csv', 'db'])
    ap.add_argument('--sl', default='none', choices=['none', 'atr'])
    ap.add_argument('--sl-mult', type=float, default=1.5,
                    help='ATR 止损倍数（实盘当前 1.5，且下限 15 美元）')
    ap.add_argument('--min-trades', type=int, default=10)
    ap.add_argument('--top', type=int, default=15)
    ap.add_argument('--xlsx', default='')
    args = ap.parse_args()

    print(f"加载 {args.tf} 数据（来源: {args.source}）...")
    df = load_db(args.tf) if args.source == 'db' else load_csv(args.tf)
    df = calc_indicators(df, from_db=(args.source == 'db'))
    print(f"  {len(df)} 根K线  {df['time'].iloc[0]} → {df['time'].iloc[-1]}")

    bars = {k: df[k].values for k in
            ('open', 'high', 'low', 'close', 'rsi', 'mfi', 'adx', 'pdi', 'ndi',
             'bb_top', 'bb_mid', 'bb_bot', 'bb_width', 'atr')}
    bars['time'] = df['time'].tolist()
    bars['bb_mid_dir'] = df['bb_mid_dir'].values

    base_kwargs = dict(time_stop=48, bb_exit=8, sl_mode=args.sl, sl_atr_mult=args.sl_mult)

    # ── 基线 v2 ──
    t_v2 = run_backtest(bars, mode='v2', **base_kwargs)
    s_v2 = summarize('v2 基线(旧6筛子)', t_v2)

    print("\n" + "=" * 96)
    print("基线 vs v3 默认参数（score_min=4, adx_max=40, 无硬门禁插针, 背离10根）")
    print("=" * 96)
    p_def = EntryParams()
    t_v3 = run_backtest(bars, mode='v3', p=p_def, **base_kwargs)
    s_v3 = summarize('v3 默认', t_v3)
    _print_table([s_v2, s_v3])

    # ── 网格扫描 ──
    grid = []
    for score_min in (3, 4, 5, 6):
        for adx_max in (30.0, 40.0, 60.0):
            for req_pierce in (False, True):
                for div_lb in (8, 10, 14):
                    grid.append(EntryParams(score_min=score_min, adx_max=adx_max,
                                            require_pierce=req_pierce, div_lookback=div_lb))
    print(f"\n扫描 {len(grid)} 组 v3 参数...")
    results = []
    for i, p in enumerate(grid):
        tr = run_backtest(bars, mode='v3', p=p, **base_kwargs)
        s = summarize(f"S{p.score_min}/A{int(p.adx_max)}/{'P' if p.require_pierce else '-'}/D{p.div_lookback}", tr)
        s['params'] = p.as_dict()
        s['trades_list'] = tr
        results.append(s)
        if (i + 1) % 20 == 0:
            print(f"  {i+1}/{len(grid)}")

    ok = [r for r in results if r['trades'] >= args.min_trades]
    ok.sort(key=lambda r: r['pnl'], reverse=True)
    print(f"\n样本 ≥ {args.min_trades} 笔的组合: {len(ok)} / {len(results)}")
    print("=" * 96)
    print(f"净 PnL Top {args.top}")
    print("=" * 96)
    _print_table(ok[:args.top])

    ok2 = sorted(ok, key=lambda r: r['avg_pnl'], reverse=True)
    print("\n" + "=" * 96)
    print(f"平均每笔盈亏 Top {args.top}（样本少时比总 PnL 更可靠）")
    print("=" * 96)
    _print_table(ok2[:args.top])

    out = {'baseline_v2': s_v2, 'v3_default': s_v3,
           'grid': [{k: v for k, v in r.items() if k != 'trades_list'} for r in results]}
    jpath = os.path.join(ROOT, 'backtest', f'fish_eaten_v3_{args.tf}_{args.source}.json')
    with open(jpath, 'w', encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False, indent=2, default=str)
    print(f"\nJSON: {jpath}")

    if args.xlsx:
        _write_xlsx(args.xlsx, s_v2, s_v3, results, t_v2, t_v3)
        print(f"Excel: {args.xlsx}")


def _print_table(rows):
    hdr = (f"{'组合':<26}{'笔数':>5}{'胜率%':>7}{'净PnL':>10}{'均笔':>9}{'回撤':>9}"
           f"{'均K线':>7}{'ADV6':>8}{'ADV12':>8}{'MFE12':>8}")
    print(hdr)
    print("-" * len(hdr))
    for r in rows:
        print(f"{r['name']:<26}{r['trades']:>5}{r['winrate']:>7.1f}{r['pnl']:>10.2f}"
              f"{r['avg_pnl']:>9.2f}{r['maxdd']:>9.2f}{r['avg_bars']:>7.1f}"
              f"{r['adv6']:>8.2f}{r['adv12']:>8.2f}{r['mfe12']:>8.2f}")
    print("\nADV6 / ADV12 = 入场后 6 / 12 根内最大逆向幅度（越小越好：不是被趋势推着走）")
    print("MFE12        = 入场后 12 根内最大有利偏移（越大越好）")


def _write_xlsx(path, s_v2, s_v3, results, t_v2, t_v3):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    hf = Font(bold=True, color='FFFFFF', size=11)
    fill = PatternFill('solid', fgColor='2F5496')

    ws = wb.create_sheet('汇总')
    cols = ['组合', '笔数', '胜率%', '净PnL', '均笔盈亏', '最大回撤', '平均K线',
            'ADV6', 'ADV12', 'MFE12', 'score_min', 'adx_max', 'require_pierce', 'div_lookback']
    for c, h in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hf
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center')
    row = 2
    for r in [s_v2, s_v3] + sorted(results, key=lambda x: x['pnl'], reverse=True):
        pr = r.get('params', {})
        vals = [r['name'], r['trades'], r['winrate'], r['pnl'], r['avg_pnl'], r['maxdd'],
                r['avg_bars'], r['adv6'], r['adv12'], r['mfe12'],
                pr.get('score_min', ''), pr.get('adx_max', ''),
                pr.get('require_pierce', ''), pr.get('div_lookback', '')]
        for c, v in enumerate(vals, 1):
            ws.cell(row=row, column=c, value=v).alignment = Alignment(horizontal='center')
        row += 1
    for c, w in enumerate([26, 6, 8, 10, 10, 10, 9, 8, 8, 8, 10, 9, 14, 12], 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    for name, trades in [('v2基线', t_v2), ('v3默认', t_v3)]:
        ds = wb.create_sheet(name)
        dcols = ['方向', '入场时间', '入场价', '出场时间', '出场价', 'K线', 'PnL', '原因',
                 'ADV6', 'ADV12', 'MFE12']
        for c, h in enumerate(dcols, 1):
            cell = ds.cell(row=1, column=c, value=h)
            cell.font = hf
            cell.fill = fill
        for i, t in enumerate(trades, 2):
            for c, v in enumerate([t['direction'], str(t['entry_time']), t['entry'],
                                   str(t['exit_time']), t['exit'], t['bars'], t['pnl'],
                                   t['reason'], t['adv6'], t['adv12'], t['mfe12']], 1):
                ds.cell(row=i, column=c, value=v).alignment = Alignment(horizontal='center')
        for c, w in enumerate([8, 20, 10, 20, 10, 6, 9, 11, 8, 8, 8], 1):
            ds.column_dimensions[get_column_letter(c)].width = w
    wb.save(path)


if __name__ == '__main__':
    main()
