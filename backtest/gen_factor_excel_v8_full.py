"""M30 RSI v8 — 5因子阶梯分 + ls-ss净得分 完整分析表
用来排查RSI值与RSI交叉互相抵消的问题"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "全组合"

hdr_font = Font(bold=True, size=11, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="2F5496")
green = PatternFill("solid", fgColor="C6EFCE")
red = PatternFill("solid", fgColor="FFC7CE")
yellow = PatternFill("solid", fgColor="FFEB9C")
orange = PatternFill("solid", fgColor="F4B183")
thin = Side(style='thin')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── 因子取值 ──
# ① MA14趋势 (用户要求改回来的)
ma_opts = {"<MA14(做空趋势)": (0,1), ">MA14(做多趋势)": (1,0)}
# ② BB
bb_opts = {"下轨": (1,0), "下轨~中轨": (0,0), "中轨~上轨": (0,0), "上轨": (0,1)}
# ③ RSI值 - 阶梯分
rsi_opts = {"<30(深超卖)": (2,0), "30-50(弱低位)": (1,0), "50-70(弱高位)": (0,1), ">70(深超买)": (0,2)}
# ④ RSI交叉 - 阶梯分
cross_opts = {"6>13>27(强金叉)": (2,0), "6>13(弱金叉)": (1,0), "6<13(弱死叉)": (0,1), "6<13<27(强死叉)": (0,2)}
# ⑤ DI强弱 - 阶梯分
di_opts = {"+DI>2×-DI(强多)": (2,0), "+DI>-DI(弱多)": (1,0), "-DI>+DI(弱空)": (0,1), "-DI>2×+DI(强空)": (0,2)}

def is_possible(rsi, cross):
    """标记理论冲突: RSI值与RSI交叉方向相反"""
    rsi_dir = 'bull' if rsi in ("<30(深超卖)", "30-50(弱低位)") else 'bear' if rsi in (">70(深超买)", "50-70(弱高位)") else 'neutral'
    cross_dir = 'bull' if cross in ("6>13>27(强金叉)", "6>13(弱金叉)") else 'bear' if cross in ("6<13(弱死叉)", "6<13<27(强死叉)") else 'neutral'
    if rsi_dir == 'bull' and cross_dir == 'bear':
        return False, f"RSI{rsi}做多 vs 交叉{cross}做空 → 抵消"
    if rsi_dir == 'bear' and cross_dir == 'bull':
        return False, f"RSI{rsi}做空 vs 交叉{cross}做多 → 抵消"
    # 强金叉时RSI不可能<30
    if cross == "6>13>27(强金叉)" and rsi in ("<30(深超卖)", "30-50(弱低位)"):
        return False, "强金叉时RSI不可能在低位"
    # 强死叉时RSI不可能>70
    if cross == "6<13<27(强死叉)" and rsi in (">70(深超买)", "50-70(弱高位)"):
        return False, "强死叉时RSI不可能在高位"
    return True, ""

def bb_possible(bb, rsi):
    if bb == "下轨" and rsi in ("50-70(弱高位)", ">70(深超买)"):
        return False
    if bb == "上轨" and rsi in ("<30(深超卖)", "30-50(弱低位)"):
        return False
    return True

HEADERS = [
    "#", "MA14", "BB", "RSI值", "RSI交叉", "DI",
    "①MA14", "②BB", "③RSI值", "④RSI交叉", "⑤DI",
    "ls", "ss", "净分(ls-ss)", "≥3?", "≥4?", "冲突说明"
]

for col, h in enumerate(HEADERS, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = Alignment(horizontal='center', wrap_text=True)
    c.border = border

rows = []
seq = 0
conflict_rows = []
ok_rows = []

for ma_lbl, (ma_l, ma_s) in ma_opts.items():
    for bb_lbl, (bb_l, bb_s) in bb_opts.items():
        for rsi_lbl, (r_l, r_s) in rsi_opts.items():
            if not bb_possible(bb_lbl, rsi_lbl):
                continue
            for cross_lbl, (c_l, c_s) in cross_opts.items():
                possible, reason = is_possible(rsi_lbl, cross_lbl)
                for di_lbl, (d_l, d_s) in di_opts.items():
                    seq += 1
                    ls = ma_l + bb_l + r_l + c_l + d_l
                    ss = ma_s + bb_s + r_s + c_s + d_s
                    net = ls - ss

                    ok3 = "✅" if net >= 3 else "❌"
                    ok4 = "✅" if net >= 4 else "❌"

                    # 冲突标记
                    conflict_note = ""
                    if not possible:
                        # 因子冲突: RSI和交叉方向相反
                        if rsi_lbl.startswith("<30") and "死叉" in cross_lbl:
                            conflict_note = "⚠️ RSI超卖喊多, 交叉死叉喊空, 互抵"
                        elif rsi_lbl.startswith(">70") and "金叉" in cross_lbl:
                            conflict_note = "⚠️ RSI超买喊空, 交叉金叉喊多, 互抵"
                        elif "强金叉" in cross_lbl and rsi_lbl in ("<30", "30-50"):
                            conflict_note = "⚠️ 不可能: 强金叉时RSI不会<50"
                        elif "强死叉" in cross_lbl and rsi_lbl in (">70", "50-70"):
                            conflict_note = "⚠️ 不可能: 强死叉时RSI不会>50"

                    row = [seq, ma_lbl, bb_lbl, rsi_lbl, cross_lbl, di_lbl,
                           ma_l or ma_s, bb_l or bb_s, r_l or r_s, c_l or c_s, d_l or d_s,
                           ls, ss, net, ok3, ok4, conflict_note]
                    rows.append(row)

# 排序: 按净分降序
rows.sort(key=lambda r: r[12], reverse=True)

for i, r in enumerate(rows, 2):
    for col, val in enumerate(r, 1):
        c = ws.cell(row=i, column=col, value=val)
        c.border = border
        c.alignment = Alignment(horizontal='center', wrap_text=True)

        h = HEADERS[col-1]
        # 净分着色
        if h == "净分(ls-ss)" and isinstance(val, int):
            if val >= 4: c.fill = green
            elif val >= 3: c.fill = orange
            elif val <= -3: c.fill = red
        # 冲突说明着色
        if h == "冲突说明" and "⚠️" in str(val):
            c.fill = yellow

# 列宽
for i, w in enumerate([4, 18, 12, 14, 18, 18, 6, 6, 6, 6, 6, 6, 6, 12, 6, 6, 40], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A2"

# ── Sheet2: 分析结论 ──
ws2 = wb.create_sheet("分析")
ws2.cell(row=1, column=1, value="v8 5因子评分问题分析").font = Font(bold=True, size=14)
ws2.merge_cells("A1:C1")

issues = [
    ("问题1: RSI值和RSI交叉互相抵消",
     "RSI值给ls分(RSI<30) + RSI交叉给ss分(死叉) = 净分≈0",
     "需要RSI值和RSI交叉方向一致才能高分"),
    ("问题2: 弱信号给分太多",
     "RSI 30-50 给+1, 弱金叉给+1, 但大部分时间RSI就在30-50之间晃",
     "这些'常亮'信号让净分很难拉开差距"),
    ("问题3: MA14趋势和DI强弱方向重复",
     "MA14>MA14(做多) + DI>+DI(做多) → 都在喊多, 重复给分",
     "牛市时两个因子都喊多, 熊市时都喊空, 不独立"),
]

for i, (title, desc, fix) in enumerate(issues, 3):
    ws2.cell(row=i, column=1, value=title).font = Font(bold=True)
    ws2.cell(row=i, column=1).fill = yellow
    ws2.cell(row=i+1, column=1, value=f"现象: {desc}")
    ws2.cell(row=i+2, column=1, value=f"对策: {fix}")

# 统计区
from collections import Counter
net_counts = Counter(r[12] for r in rows)
total = len(rows)
ws2.cell(row=10, column=1, value="").font = hdr_font
ws2.cell(row=11, column=1, value="净分分布").font = Font(bold=True)
ws2.cell(row=12, column=1, value="净分").font = hdr_font
ws2.cell(row=12, column=1).fill = hdr_fill
ws2.cell(row=12, column=2, value="组合数").font = hdr_font
ws2.cell(row=12, column=2).fill = hdr_fill

row_i = 13
for net_val in sorted(net_counts.keys()):
    cnt = net_counts[net_val]
    ws2.cell(row=row_i, column=1, value=net_val).border = border
    ws2.cell(row=row_i, column=2, value=cnt).border = border
    if net_val >= 3: ws2.cell(row=row_i, column=1).fill = green
    elif net_val <= -3: ws2.cell(row=row_i, column=1).fill = red
    row_i += 1

ws2.cell(row=row_i, column=1, value=f"共{total}组合").font = hdr_font
row_i += 2
ws2.cell(row=row_i, column=1, value=f"净分≥3(做多): {sum(1 for r in rows if r[12]>=3)}")
row_i += 1
ws2.cell(row=row_i, column=1, value=f"净分≤-3(做空): {sum(1 for r in rows if r[12]<=-3)}")
row_i += 1
ws2.cell(row=row_i, column=1, value=f"冲突抵消(黄色标记): {sum(1 for r in rows if '⚠️' in str(r[16]))}")

# 特殊标注: 列出净分≥4的场景
row_i += 2
ws2.cell(row=row_i, column=1, value="净分≥4 场景(真正能进场的)").font = Font(bold=True, color="006100")
row_i += 1
ws2.cell(row=row_i, column=1, value="MA14").font = hdr_font
ws2.cell(row=row_i, column=1).fill = hdr_fill
ws2.cell(row=row_i, column=2, value="BB").font = hdr_font
ws2.cell(row=row_i, column=2).fill = hdr_fill
ws2.cell(row=row_i, column=3, value="RSI值").font = hdr_font
ws2.cell(row=row_i, column=3).fill = hdr_fill
ws2.cell(row=row_i, column=4, value="RSI交叉").font = hdr_font
ws2.cell(row=row_i, column=4).fill = hdr_fill
ws2.cell(row=row_i, column=5, value="DI").font = hdr_font
ws2.cell(row=row_i, column=5).fill = hdr_fill
ws2.cell(row=row_i, column=6, value="净分").font = hdr_font
ws2.cell(row=row_i, column=6).fill = hdr_fill
row_i += 1

for r in rows:
    if r[12] >= 4:
        for j in [1,2,3,4,5,12]:
            ws2.cell(row=row_i, column=j if j<6 else 6, value=r[j])
            ws2.cell(row=row_i, column=j if j<6 else 6).border = border
        row_i += 1

wb.save("D:/backup/BaoBao/PythonProgram/xauusd/backtest/factor_v8_full.xlsx")
print(f"Done! {len(rows)} combos, net≥3: {sum(1 for r in rows if r[12]>=3)}, net≥4: {sum(1 for r in rows if r[12]>=4)}")
print(f"Conflict rows: {sum(1 for r in rows if '⚠️' in str(r[16]))}")
