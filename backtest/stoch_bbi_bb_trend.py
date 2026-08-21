"""
Stoch+BBI+BB 趋势跟踪 v2 · ±DI 门禁代替 ADX
TA-Lib 计算，逐根 K 线判定
"""
import os, sys, numpy as np, pandas as pd, talib
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

LOT = 0.01
DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'data')

def load_data(tf):
    path = os.path.join(DATA_DIR, f'XAUUSD_{tf}.csv')
    df = pd.read_csv(path, names=['time','open','high','low','close','volume'], skiprows=1, parse_dates=['time'])
    return df

def calc_indicators(df, stoch_k=5, stoch_d=3):
    o, h, l, c, v = df['open'].values, df['high'].values, df['low'].values, df['close'].values, df['volume'].values
    
    # BBI = (SMA3 + SMA6 + SMA12 + SMA24) / 4
    sma3 = talib.SMA(c, timeperiod=3)
    sma6 = talib.SMA(c, timeperiod=6)
    sma12 = talib.SMA(c, timeperiod=12)
    sma24 = talib.SMA(c, timeperiod=24)
    df['bbi'] = (sma3 + sma6 + sma12 + sma24) / 4
    
    # Stoch
    slowk, slowd = talib.STOCH(h, l, c, fastk_period=stoch_k, slowk_period=stoch_d, slowk_matype=0, slowd_period=3, slowd_matype=0)
    df['stoch_k'] = slowk
    df['stoch_d'] = slowd
    
    # BB
    bb_up, bb_mid, bb_low = talib.BBANDS(c, timeperiod=20, nbdevup=2, nbdevdn=2)
    df['bb_top'] = bb_up
    df['bb_mid'] = bb_mid
    df['bb_bot'] = bb_low
    
    # ±DI (14)
    df['pdi'] = talib.PLUS_DI(h, l, c, timeperiod=14)
    df['ndi'] = talib.MINUS_DI(h, l, c, timeperiod=14)
    
    # BBI 斜率
    df['bbi_slope'] = np.where(df['bbi'].diff() > 0, 'up', np.where(df['bbi'].diff() < 0, 'down', 'flat'))
    
    # ATR
    df['atr'] = talib.ATR(h, l, c, timeperiod=14)
    
    return df

def run_backtest(df, params):
    stoch_k_period = params['stoch_k']
    exit_confirm_bars = params['exit_confirm']
    di_gate = params['di_gate']
    trail_atr = params.get('trail_atr', 0)  # 0 = 不使用 trailing stop
    
    df = calc_indicators(df, stoch_k=stoch_k_period)
    
    trades = []
    in_position = False
    direction = None
    entry_price = 0.0
    entry_time = None
    entry_bar = -1
    exit_count = 0
    trail_peak = 0.0  # trailing stop 跟踪的最高价(多头)/最低价(空头)
    pending_order = None
    
    for i in range(len(df)):
        row = df.iloc[i]
        open_p = row['open']
        close_p = row['close']
        bbi = row['bbi']
        stoch_k = row['stoch_k']
        stoch_d = row['stoch_d']
        bb_top = row['bb_top']
        bb_mid = row['bb_mid']
        bb_bot = row['bb_bot']
        bbi_slope = row['bbi_slope']
        pdi = row['pdi']
        ndi = row['ndi']
        
        if np.isnan(bbi) or np.isnan(stoch_k) or np.isnan(pdi):
            continue
        
        # 执行挂单
        if pending_order == 'buy' and not in_position:
            entry_price = open_p
            entry_time = row['time']
            direction = 'LONG'
            in_position = True
            entry_bar = i
            exit_count = 0
            pending_order = None
            continue
        elif pending_order == 'sell' and not in_position:
            entry_price = open_p
            entry_time = row['time']
            direction = 'SHORT'
            in_position = True
            entry_bar = i
            exit_count = 0
            pending_order = None
            continue
        
        # ── 持仓中：出场 ──
        if in_position:
            bars_held = i - entry_bar
            
            if direction == 'LONG':
                # 跟踪最高价
                trail_peak = max(trail_peak, row['high'])
                
                # 硬止损：跌破 BB 下轨
                if close_p < bb_bot:
                    pnl = close_p - entry_price
                    trades.append(dict(entry_time=entry_time, exit_time=row['time'], direction='LONG',
                        entry_price=round(entry_price,2), exit_price=round(close_p,2),
                        bars=bars_held, pnl=round(pnl,2), reason='bb_stop'))
                    in_position = False; direction = None; continue
                
                # Trailing stop：从最高点回撤 ATR 倍数
                if trail_atr > 0 and close_p < trail_peak - trail_atr * row['atr']:
                    pnl = close_p - entry_price
                    trades.append(dict(entry_time=entry_time, exit_time=row['time'], direction='LONG',
                        entry_price=round(entry_price,2), exit_price=round(close_p,2),
                        bars=bars_held, pnl=round(pnl,2), reason='trailing_stop'))
                    in_position = False; direction = None; continue
                
                # 趋势反转出场：连续 N 根 < BBI + BBI 斜率向下
                if close_p < bbi:
                    exit_count += 1
                else:
                    exit_count = 0
                
                if exit_count >= exit_confirm_bars and bbi_slope == 'down':
                    pnl = close_p - entry_price
                    trades.append(dict(entry_time=entry_time, exit_time=row['time'], direction='LONG',
                        entry_price=round(entry_price,2), exit_price=round(close_p,2),
                        bars=bars_held, pnl=round(pnl,2), reason='trend_reversal'))
                    in_position = False; direction = None; continue
                
            else:  # SHORT
                trail_peak = min(trail_peak, row['low']) if trail_peak != 0 else row['low']
                
                if close_p > bb_top:
                    pnl = entry_price - close_p
                    trades.append(dict(entry_time=entry_time, exit_time=row['time'], direction='SHORT',
                        entry_price=round(entry_price,2), exit_price=round(close_p,2),
                        bars=bars_held, pnl=round(pnl,2), reason='bb_stop'))
                    in_position = False; direction = None; continue
                
                if trail_atr > 0 and close_p > trail_peak + trail_atr * row['atr']:
                    pnl = entry_price - close_p
                    trades.append(dict(entry_time=entry_time, exit_time=row['time'], direction='SHORT',
                        entry_price=round(entry_price,2), exit_price=round(close_p,2),
                        bars=bars_held, pnl=round(pnl,2), reason='trailing_stop'))
                    in_position = False; direction = None; continue
                
                if close_p > bbi:
                    exit_count += 1
                else:
                    exit_count = 0
                
                if exit_count >= exit_confirm_bars and bbi_slope == 'up':
                    pnl = entry_price - close_p
                    trades.append(dict(entry_time=entry_time, exit_time=row['time'], direction='SHORT',
                        entry_price=round(entry_price,2), exit_price=round(close_p,2),
                        bars=bars_held, pnl=round(pnl,2), reason='trend_reversal'))
                    in_position = False; direction = None; continue
            
            continue
        
        # ── 空仓：入场 ──
        # ±DI 门禁：趋势方向确认
        if di_gate > 0 and abs(pdi - ndi) <= di_gate:
            continue
        
        # 多头：+DI > -DI (趋势向上), 价格>BBI, Stoch金叉, K<80, 价格>=BB中轨
        if pdi > ndi and close_p > bbi and stoch_k > stoch_d and stoch_k < 80 and close_p >= bb_mid:
            pending_order = 'buy'
        # 空头：-DI > +DI (趋势向下), 价格<BBI, Stoch死叉, K>20, 价格<=BB中轨
        elif ndi > pdi and close_p < bbi and stoch_k < stoch_d and stoch_k > 20 and close_p <= bb_mid:
            pending_order = 'sell'
    
    return trades


def main():
    timeframes = ['M15', 'M30']
    param_grid = []
    for stoch_k in [5]:
        for exit_cfm in [3]:
            for di_gate in [5]:
                for trail_atr in [0, 1.5, 2.0, 3.0]:
                    for tf in timeframes:
                        param_grid.append({'stoch_k': stoch_k, 'exit_confirm': exit_cfm, 'di_gate': di_gate, 'trail_atr': trail_atr, 'timeframe': tf})
    
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='2F5496')
    green_fill = PatternFill('solid', fgColor='E2EFDA')
    red_fill = PatternFill('solid', fgColor='FCE4EC')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'), bottom=Side(style='thin'))
    
    wb = openpyxl.Workbook()
    if 'Sheet' in wb.sheetnames: del wb['Sheet']
    
    all_results = []
    all_trades = {}
    
    for p in param_grid:
        tf = p['timeframe']
        df = load_data(tf)
        combo = f"{tf}_Stoch{p['stoch_k']}_Cfm{p['exit_confirm']}_DI{p['di_gate']}_Trail{p['trail_atr']}"
        print(f'[{combo}] {len(df)}根K线...', end=' ')
        
        trades = run_backtest(df, p)
        total = len(trades)
        wins = sum(1 for t in trades if t['pnl'] > 0)
        losses = sum(1 for t in trades if t['pnl'] <= 0)
        total_pnl = sum(t['pnl'] for t in trades)
        winrate = wins / total * 100 if total else 0
        max_win = max((t['pnl'] for t in trades), default=0)
        max_loss = min((t['pnl'] for t in trades), default=0)
        avg_win = sum(t['pnl'] for t in trades if t['pnl'] > 0) / wins if wins else 0
        avg_loss = sum(t['pnl'] for t in trades if t['pnl'] <= 0) / losses if losses else 0
        profit_factor = abs(avg_win / avg_loss) if avg_loss != 0 else 0
        
        r = dict(combo=combo, tf=tf, stoch_k=p['stoch_k'], exit_cfm=p['exit_confirm'], di_gate=p['di_gate'], trail_atr=p['trail_atr'],
                 pnl=round(total_pnl,2), trades=total, wins=wins, losses=losses, winrate=round(winrate,1),
                 max_win=round(max_win,2), max_loss=round(max_loss,2), avg_pnl=round(total_pnl/total,2) if total else 0,
                 profit_factor=round(profit_factor,2))
        all_results.append(r)
        all_trades[combo] = trades
        print(f'{total}笔 | 胜率{winrate:.0f}% | PnL:${total_pnl:+6.2f} | 盈亏比:{profit_factor:.2f}')
    
    all_results.sort(key=lambda r: r['pnl'], reverse=True)
    
    # 汇总
    ws = wb.create_sheet(title='汇总')
    ws.merge_cells('A1:L1')
    ws['A1'] = 'Stoch+BBI+BB 趋势跟踪 v2 · ±DI 门禁 · 回测'
    ws['A1'].font = Font(bold=True, size=14, color='2F5496')
    ws.merge_cells('A2:L2')
    ws['A2'] = '入场: BBI+Stoch金叉+±DI方向+BB中轨 | 出场: 连续N根<BBI+BBI斜率向下 | 硬止损: BB反向轨'
    ws['A2'].font = Font(italic=True, color='666666')
    
    headers = ['组合', '周期', 'Stoch', '确认', 'DI门禁', 'Trail', '净PnL($)', '交易', '胜率(%)', '胜笔', '亏笔', '盈亏比', '平均PnL']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center'); cell.border = thin_border
    
    for i, r in enumerate(all_results, 5):
        vals = [r['combo'], r['tf'], r['stoch_k'], r['exit_cfm'], r['di_gate'], r['trail_atr'],
                r['pnl'], r['trades'], r['winrate'], r['wins'], r['losses'], r['profit_factor'], r['avg_pnl']]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=v)
            cell.border = thin_border; cell.alignment = Alignment(horizontal='center')
            if col == 7:
                if v > 0: cell.fill = green_fill; cell.font = Font(color='006100')
                elif v < 0: cell.fill = red_fill; cell.font = Font(color='9C0006')
    
    widths = [38, 8, 8, 8, 10, 8, 14, 8, 10, 8, 8, 10, 10]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    for col in range(1, 14):
        ws.cell(row=5, column=col).font = Font(bold=True, color='B8860B')
    
    # 明细
    da = wb.create_sheet(title='全部明细')
    dheaders = ['组合', '方向', '入场时间', '入场价', '出场时间', '出场价', '持仓K线', 'PnL($)', '出场原因']
    for col, h in enumerate(dheaders, 1):
        cell = da.cell(row=1, column=col, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center'); cell.border = thin_border
    
    dr = 2
    for r in all_results:
        for t in all_trades.get(r['combo'], []):
            vals = [r['combo'], t['direction'], t['entry_time'].strftime('%Y-%m-%d %H:%M'),
                    t['entry_price'], t['exit_time'].strftime('%Y-%m-%d %H:%M'), t['exit_price'],
                    t['bars'], t['pnl'], t['reason']]
            for col, v in enumerate(vals, 1):
                cell = da.cell(row=dr, column=col, value=v)
                cell.border = thin_border; cell.alignment = Alignment(horizontal='center')
                if col == 8:
                    if v > 0: cell.fill = green_fill
                    elif v < 0: cell.fill = red_fill
            dr += 1
    da.auto_filter.ref = f'A1:I{dr-1}'
    dwidths = [32, 10, 18, 12, 18, 12, 12, 12, 14]
    for col, w in enumerate(dwidths, 1):
        da.column_dimensions[get_column_letter(col)].width = w
    
    out_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                            'backtest', 'stoch_bbi_bb_trend_v3_trail.xlsx')
    wb.save(out_path)
    print(f'\n✅ Excel: {out_path}')


if __name__ == '__main__':
    main()