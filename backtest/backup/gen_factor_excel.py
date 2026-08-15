"""生成 M30 RSI 因子评分 + DI门禁 完整分析表"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "因子评分分析"

# ── 样式 ──
hdr_font = Font(bold=True, size=11, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="2F5496")
green_fill = PatternFill("solid", fgColor="C6EFCE")
red_fill = PatternFill("solid", fgColor="FFC7CE")
yellow_fill = PatternFill("solid", fgColor="FFEB9C")
light_fill = PatternFill("solid", fgColor="F2F2F2")
thin = Side(style='thin')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

HEADERS = [
    "BB位置", "RSI范围", "RSIdir", "MA14位置",
    "①MA14", "②BB", "③RSI", "④RSIdir",
    "ls总分", "ss总分", "ls≥3?", "ss≥3?",
    "+DI>-DI结论", "-DI>+DI结论", "说明"
]

for col, h in enumerate(HEADERS, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = Alignment(horizontal='center', wrap_text=True)
    c.border = border

# ── 数据 ──
bb_positions = ["下轨", "下轨~中轨", "中轨~上轨", "上轨"]
rsi_ranges = ["<30", "30-50", "50-70", ">70"]  # index 0,1,2,3
rsi_dirs = ["↑向上", "↓向下"]
ma14_positions = ["↑上方", "↓下方"]

def bb_score(bb_pos):
    if bb_pos == "下轨": return (1, 0)
    if bb_pos == "上轨": return (0, 1)
    return (0, 0)

def rsi_score(rsi_range):
    # <30 → long=1; >70 → short=1; 中间 → 0
    if rsi_range == "<30": return (1, 0)
    if rsi_range in ("50-70", ">70"):
        # 按代码 >65 才 short=1, 这里用 >70 做边界
        return (0, 0) if rsi_range == "50-70" else (0, 1)
    return (0, 0)

def rsi_dir_score(rsi_dir, side):
    # side: 'long' → 向上+1, 'short' → 向下+1
    if side == 'long' and rsi_dir == "↑向上": return 1
    if side == 'short' and rsi_dir == "↓向下": return 1
    return 0

def is_possible(bb_pos, rsi_range):
    """排除不可能的组合"""
    impossible = [
        ("下轨", "50-70"), ("下轨", ">70"),
        ("上轨", "<30"), ("上轨", "30-50"),
    ]
    if (bb_pos, rsi_range) in impossible:
        return False
    # 中轨两侧的极端值很罕见但标记为可能
    return True

rows = []
for bb in bb_positions:
    for rsi_r in rsi_ranges:
        if not is_possible(bb, rsi_r):
            continue
        bb_l, bb_s = bb_score(bb)
        rsi_l, rsi_s = rsi_score(rsi_r)
        for rd in rsi_dirs:
            rd_l = rsi_dir_score(rd, 'long')
            rd_s = rsi_dir_score(rd, 'short')
            for ma in ma14_positions:
                ma_l = 1 if ma == "↑上方" else 0
                ma_s = 1 if ma == "↓下方" else 0

                ls = ma_l + bb_l + rsi_l + rd_l
                ss = ma_s + bb_s + rsi_s + rd_s

                # DI门禁结论
                # +DI > -DI: only allow LONG (ls≥3)
                # -DI > +DI: only allow SHORT (ss≥3)
                if ls >= 3:
                    di_up = "✅ 做多"
                else:
                    di_up = "❌ 分不够"
                if ss >= 3:
                    di_down = "✅ 做空"
                else:
                    di_down = "❌ 分不够"

                # 特殊: 分够但被门禁拦截
                if ls >= 3 and ss < 3:
                    di_down = "❌ 门禁拦截"
                if ss >= 3 and ls < 3:
                    di_up = "❌ 门禁拦截"

                ls_ok = "✅" if ls >= 3 else "❌"
                ss_ok = "✅" if ss >= 3 else "❌"

                # 说明
                note = ""
                if ls >= 3 and ss >= 3:
                    note = "多空都够分,罕见"
                elif ls >= 3 and bb_l and rsi_l:
                    note = "超卖反弹"
                elif ls >= 3 and bb_l and not rsi_l:
                    note = "BB下轨支撑做多"
                elif ls >= 3 and not bb_l and rsi_l:
                    note = "RSI超卖做多"
                elif ss >= 3 and bb_s and rsi_s:
                    note = "超买回落做空"
                elif ss >= 3 and bb_s and not rsi_s:
                    note = "BB上轨压力做空"
                elif ss >= 3 and not bb_s and rsi_s:
                    note = "RSI超买做空"

                r = [
                    bb, rsi_r, rd, ma,
                    ma_l, bb_l if bb_l else (bb_s if bb_s else 0), rsi_l if rsi_l else (rsi_s if rsi_s else 0), rd_l if rd_l else (rd_s if rd_s else 0),
                    ls, ss, ls_ok, ss_ok,
                    di_up, di_down, note
                ]
                rows.append(r)

# 排序: BB位置 → RSI范围 → RSI方向 → MA14
def sort_key(r):
    bb_order = {"下轨":0, "下轨~中轨":1, "中轨~上轨":2, "上轨":3}
    rsi_order = {"<30":0, "30-50":1, "50-70":2, ">70":3}
    rd_order = {"↑向上":0, "↓向下":1}
    ma_order = {"↑上方":0, "↓下方":1}
    return (bb_order.get(r[0],9), rsi_order.get(r[1],9), rd_order.get(r[2],9), ma_order.get(r[3],9))

rows.sort(key=sort_key)

for i, r in enumerate(rows, 2):
    for col, val in enumerate(r, 1):
        c = ws.cell(row=i, column=col, value=val)
        c.border = border
        c.alignment = Alignment(horizontal='center', wrap_text=True)

        # 条件着色
        col_name = HEADERS[col-1]
        if col_name == "ls总分" and isinstance(val, int) and val >= 3:
            c.fill = green_fill
        elif col_name == "ss总分" and isinstance(val, int) and val >= 3:
            c.fill = green_fill
        elif col_name in ("ls≥3?", "ss≥3?") and val == "✅":
            c.fill = green_fill
        elif val in ("❌ 门禁拦截",):
            c.fill = red_fill
        elif val in ("❌ 分不够",):
            c.fill = yellow_fill

# ── 列宽 ──
widths = [12, 8, 8, 10, 8, 6, 6, 8, 8, 8, 8, 8, 14, 14, 20]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── 冻结首行 ──
ws.freeze_panes = "A2"

# ── Sheet2: 汇总 ──
ws2 = wb.create_sheet("汇总")
ws2.cell(row=1, column=1, value="场景").font = hdr_font
ws2.cell(row=1, column=1).fill = hdr_fill
ws2.cell(row=1, column=2, value="BB位置").font = hdr_font
ws2.cell(row=1, column=2).fill = hdr_fill
ws2.cell(row=1, column=3, value="RSI").font = hdr_font
ws2.cell(row=1, column=3).fill = hdr_fill
ws2.cell(row=1, column=4, value="RSIdir").font = hdr_font
ws2.cell(row=1, column=4).fill = hdr_fill
ws2.cell(row=1, column=5, value="MA14").font = hdr_font
ws2.cell(row=1, column=5).fill = hdr_fill
ws2.cell(row=1, column=6, value="评分").font = hdr_font
ws2.cell(row=1, column=6).fill = hdr_fill
ws2.cell(row=1, column=7, value="+DI>-DI(升势)").font = hdr_font
ws2.cell(row=1, column=7).fill = hdr_fill
ws2.cell(row=1, column=8, value="-DI>+DI(降势)").font = hdr_font
ws2.cell(row=1, column=8).fill = hdr_fill

summary = [r for r in rows if (isinstance(r[8], int) and r[8] >= 3) or (isinstance(r[9], int) and r[9] >= 3)]
for i, r in enumerate(summary, 2):
    ws2.cell(row=i, column=1, value=f"场景{i-1}").border = border
    ws2.cell(row=i, column=2, value=r[0]).border = border  # BB
    ws2.cell(row=i, column=3, value=r[1]).border = border  # RSI
    ws2.cell(row=i, column=4, value=r[2]).border = border  # RSIdir
    ws2.cell(row=i, column=5, value=r[3]).border = border  # MA14
    # 评分文字
    if r[8] >= 3 and r[9] < 3:
        ws2.cell(row=i, column=6, value=f"ls={r[8]}").border = border
    elif r[9] >= 3 and r[8] < 3:
        ws2.cell(row=i, column=6, value=f"ss={r[9]}").border = border
    else:
        ws2.cell(row=i, column=6, value=f"ls={r[8]}/ss={r[9]}").border = border
    ws2.cell(row=i, column=7, value=r[12]).border = border  # +DI结论
    ws2.cell(row=i, column=8, value=r[13]).border = border  # -DI结论

    # 着色
    if "✅" in str(r[12]) and "❌ 门禁" not in str(r[12]):
        ws2.cell(row=i, column=7).fill = green_fill
    if "✅" in str(r[13]) and "❌ 门禁" not in str(r[13]):
        ws2.cell(row=i, column=8).fill = green_fill
    if "❌ 门禁" in str(r[12]) or "❌ 门禁" in str(r[13]):
        ws2.cell(row=i, column=6).fill = red_fill
    if "✅" in str(r[12]) and "✅" in str(r[13]):
        ws2.cell(row=i, column=6).fill = yellow_fill

for i, w in enumerate([10, 12, 8, 8, 10, 12, 16, 16], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

wb.save("D:/backup/BaoBao/PythonProgram/xauusd/backtest/factor_table.xlsx")
print(f"Done! {len(rows)} rows written to factor_table.xlsx")
print(f"Summary: {len(summary)} tradeable scenarios")
