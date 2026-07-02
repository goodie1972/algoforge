"""M30 RSI — 5因子阶梯分 + ls-ss净得分 ≥+4做多 / ≤-4做空"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "全组合"

hdr_font = Font(bold=True, size=11, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="2F5496")
green = PatternFill("solid", fgColor="C6EFCE")
red = PatternFill("solid", fgColor="FFC7CE")
yellow = PatternFill("solid", fgColor="FFEB9C")
orange = PatternFill("solid", fgColor="F4B183")
thin = Side(style='thin')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# 因子取值
ma_opts = {"<MA14(做空信号)": (0,1), ">MA14(做多信号)": (1,0)}
bb_opts = {"下轨": (1,0), "下轨~中轨": (0,0), "中轨~上轨": (0,0), "上轨": (0,1)}
rsi_opts = {"<30(深超卖)": (2,0), "30-50(弱低位)": (1,0), "50-70(弱高位)": (0,1), ">70(深超买)": (0,2)}
cross_opts = {"6>13>27(强金叉)": (2,0), "6>13(弱金叉)": (1,0), "6<13(弱死叉)": (0,1), "6<13<27(强死叉)": (0,2)}
di_opts = {"+DI>2×-DI(强多)": (2,0), "+DI>-DI(弱多)": (1,0), "-DI>+DI(弱空)": (0,1), "-DI>2×+DI(强空)": (0,2)}

def bb_ok(bb, rsi):
    if bb == "下轨" and rsi in ("50-70(弱高位)", ">70(深超买)"): return False
    if bb == "上轨" and rsi in ("<30(深超卖)", "30-50(弱低位)"): return False
    return True

HEADERS = [
    "#", "MA14", "BB", "RSI值", "RSI交叉", "DI",
    "①MA14", "②BB", "③RSI值", "④RSI交叉", "⑤DI",
    "ls", "ss", "净分(ls-ss)", "≥+4?", "≤-4?", "信号方向"
]

for col, h in enumerate(HEADERS, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = hdr_font; c.fill = hdr_fill
    c.alignment = Alignment(horizontal='center', wrap_text=True)
    c.border = border

rows = []
seq = 0
neg_found = []
for ma_lbl, (ma_l, ma_s) in ma_opts.items():
    for bb_lbl, (bb_l, bb_s) in bb_opts.items():
        for rsi_lbl, (r_l, r_s) in rsi_opts.items():
            if not bb_ok(bb_lbl, rsi_lbl): continue
            for cross_lbl, (c_l, c_s) in cross_opts.items():
                for di_lbl, (d_l, d_s) in di_opts.items():
                    seq += 1
                    ls = ma_l + bb_l + r_l + c_l + d_l
                    ss = ma_s + bb_s + r_s + c_s + d_s
                    net = ls - ss

                    if net < 0 and len(neg_found) < 3:
                        neg_found.append((seq, ma_lbl, bb_lbl, rsi_lbl, cross_lbl, di_lbl, ls, ss, net))

                    if net >= 4: sig = "做多 ✅"
                    elif net <= -4: sig = "做空 ✅"
                    else: sig = "观望 ❌"

                    ok4_long = "✅" if net >= 4 else ""
                    ok4_short = "✅" if net <= -4 else ""

                    rows.append([seq, ma_lbl, bb_lbl, rsi_lbl, cross_lbl, di_lbl,
                                ma_l or ma_s, bb_l or bb_s, r_l or r_s, c_l or c_s, d_l or d_s,
                                ls, ss, net, ok4_long, ok4_short, sig])

nets = [r[13] for r in rows]
print(f"All unique net values: {sorted(set(nets))}")
print(f"Net >= 4 count: {sum(1 for n in nets if n >= 4)}")
print(f"Net <= -4 count: {sum(1 for n in nets if n <= -4)}")
print(f"Total rows: {len(rows)}, neg rows: {sum(1 for n in nets if n < 0)}")
if all(n >= 0 for n in nets):
    # Force generate a specific row to confirm the math
    for ma_lbl, (ma_l, ma_s) in ma_opts.items():
        for bb_lbl, (bb_l, bb_s) in bb_opts.items():
            for rsi_lbl, (r_l, r_s) in rsi_opts.items():
                for cross_lbl, (c_l, c_s) in cross_opts.items():
                    for di_lbl, (d_l, d_s) in di_opts.items():
                        ls = ma_l + bb_l + r_l + c_l + d_l
                        ss = ma_s + bb_s + r_s + c_s + d_s
                        net = ls - ss
                        if net < 0:
                            print(f"  FOUND NEGATIVE: net={net} ma={ma_lbl} bb={bb_lbl} rsi={rsi_lbl} cross={cross_lbl} di={di_lbl}")
                            print(f"    ma_l={ma_l} ma_s={ma_s} bb_l={bb_l} bb_s={bb_s} r_l={r_l} r_s={r_s} c_l={c_l} c_s={c_s} d_l={d_l} d_s={d_s}")

rows.sort(key=lambda r: r[13], reverse=True)

for i, r in enumerate(rows, 2):
    for col, val in enumerate(r, 1):
        c = ws.cell(row=i, column=col, value=val)
        c.border = border
        c.alignment = Alignment(horizontal='center', wrap_text=True)
        h = HEADERS[col-1]
        if h == "净分(ls-ss)" and isinstance(val, int):
            if val >= 4: c.fill = green
            elif val <= -4: c.fill = red
        if h == "信号方向":
            if "做多" in str(val): c.fill = green
            elif "做空" in str(val): c.fill = red
            else: c.fill = yellow

for i, w in enumerate([4, 18, 12, 14, 18, 18, 6, 6, 6, 6, 6, 6, 6, 12, 6, 6, 14], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"

# ── Sheet2: 评分规则 ──
ws2 = wb.create_sheet("评分规则")
rules = [
    ("因子", "条件", "ls加分", "ss加分"),
    ("① MA14趋势", "close > MA14 (价格在均线上方)", "1", "0"),
    ("", "close < MA14 (价格在均线下方)", "0", "1"),
    ("② BB通道", "碰下轨", "1", "0"),
    ("", "碰上轨", "0", "1"),
    ("③ RSI值", "< 30 (深度超卖)", "2", "0"),
    ("", "30 ~ 50 (偏低位)", "1", "0"),
    ("", "50 ~ 70 (偏高)", "0", "1"),
    ("", "> 70 (深度超买)", "0", "2"),
    ("④ RSI交叉(6/13/27)", "6 > 13 > 27 (强金叉)", "2", "0"),
    ("", "6 > 13 (弱金叉)", "1", "0"),
    ("", "6 < 13 (弱死叉)", "0", "1"),
    ("", "6 < 13 < 27 (强死叉)", "0", "2"),
    ("⑤ DI强弱", "+DI > 2 × (-DI) (强多)", "2", "0"),
    ("", "+DI > -DI (弱多)", "1", "0"),
    ("", "-DI > +DI (弱空)", "0", "1"),
    ("", "-DI > 2 × (+DI) (强空)", "0", "2"),
    ("", "", "", ""),
    ("进场规则", "ls - ss ≥ +4 → 做多", "", ""),
    ("", "ls - ss ≤ -4 → 做空", "", ""),
    ("", "否则观望", "", ""),
]
for i, r in enumerate(rules, 1):
    for j, v in enumerate(r, 1):
        c = ws2.cell(row=i, column=j, value=v)
        c.border = border; c.alignment = Alignment(horizontal='center')
        if i == 1: c.font = hdr_font; c.fill = hdr_fill

ws2.column_dimensions['A'].width = 20; ws2.column_dimensions['B'].width = 30

# ── Sheet3: 只显示可进场场景 ──
ws3 = wb.create_sheet("可进场(净分≥±4)")
hd3 = ["MA14", "BB", "RSI值", "RSI交叉", "DI", "净分", "信号"]
for col, h in enumerate(hd3, 1):
    c = ws3.cell(row=1, column=col, value=h)
    c.font = hdr_font; c.fill = hdr_fill
    c.alignment = Alignment(horizontal='center', wrap_text=True); c.border = border

tradeable = [r for r in rows if abs(r[13]) >= 4]
for i, r in enumerate(tradeable, 2):
    for j, idx in enumerate([1,2,3,4,5,12,16], 1):
        c = ws3.cell(row=i, column=j, value=r[idx])
        c.border = border; c.alignment = Alignment(horizontal='center')
        if j == 7 and "做多" in str(r[idx]): c.fill = green
        if j == 7 and "做空" in str(r[idx]): c.fill = red
        if j == 6 and isinstance(r[idx], int) and abs(r[idx]) >= 6: c.fill = green

for i, w in enumerate([18, 12, 14, 18, 18, 8, 12], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.freeze_panes = "A2"

# ── Sheet4: 统计 ──
ws4 = wb.create_sheet("统计")
from collections import Counter
net_counts = Counter(r[13] for r in rows)
total = len(rows)
ws4.cell(row=1, column=1, value="净分分布").font = hdr_font
ws4.cell(row=2, column=1, value="净分").font = hdr_font; ws4.cell(row=2, column=1).fill = hdr_fill
ws4.cell(row=2, column=2, value="组合数").font = hdr_font; ws4.cell(row=2, column=2).fill = hdr_fill
ws4.cell(row=2, column=3, value="占比").font = hdr_font; ws4.cell(row=2, column=3).fill = hdr_fill
row_i = 3
for n in sorted(net_counts.keys()):
    cnt = net_counts[n]
    ws4.cell(row=row_i, column=1, value=n).border = border
    ws4.cell(row=row_i, column=2, value=cnt).border = border
    ws4.cell(row=row_i, column=3, value=f"{cnt/total*100:.1f}%").border = border
    if n >= 4: ws4.cell(row=row_i, column=1).fill = green
    elif n <= -4: ws4.cell(row=row_i, column=1).fill = red
    row_i += 1

ws4.cell(row=row_i, column=1, value="合计").font = hdr_font
ws4.cell(row=row_i, column=2, value=total).font = hdr_font

buy = sum(1 for r in rows if r[13] >= 4)
sell = sum(1 for r in rows if r[13] <= -4)
row_i += 2
ws4.cell(row=row_i, column=1, value=f"做多(净分≥+4): {buy}  ({buy/total*100:.1f}%)").font = Font(bold=True, color="006100")
row_i += 1
ws4.cell(row=row_i, column=1, value=f"做空(净分≤-4): {sell}  ({sell/total*100:.1f}%)").font = Font(bold=True, color="9C0006")
row_i += 1
ws4.cell(row=row_i, column=1, value=f"观望: {total-buy-sell}  ({(total-buy-sell)/total*100:.1f}%)")

for i, w in enumerate([12, 10, 10], 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

# Debug: verify count
_debug_buy = sum(1 for r in rows if r[13] >= 4)
_debug_sell = sum(1 for r in rows if r[13] <= -4)
print(f"DEBUG: buy={_debug_buy} sell={_debug_sell} min_net={min(r[13] for r in rows)} max_net={max(r[13] for r in rows)}")

wb.save("D:/backup/BaoBao/PythonProgram/xauusd/backtest/factor_table_v8_final.xlsx")
print(f"OK! {total} combos, 做多(net>=+4): {buy}, 做空(net<=-4): {sell}, 观望: {total-buy-sell}")
