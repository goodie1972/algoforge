"""
RSI-BB Trend 价格回归 · 纯 Python 回测
不依赖 backtrader，用 TA-Lib 算指标，逐根 K 线判定
"""
import os, sys, numpy as np, pandas as pd, talib
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

LOT = 0.01  # 固定手数
DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'data')

def load_data(tf):
    """加载 CSV，返回 DataFrame"""
    path = os.path.join(DATA_DIR, f'XAUUSD_{tf}.csv')
    df = pd.read_csv(path, names=['time','open','high','low','close','volume'], skiprows=1,
                     parse_dates=['time'])
    return df

def calc_indicators(df):
    """用 TA-Lib 计算全部指标"""
    o, h, l, c, v = df['open'].values, df['high'].values, df['low'].values, df['close'].values, df['volume'].values
    
    df['rsi'] = talib.RSI(c, timeperiod=14)
    df['mfi'] = talib.MFI(h, l, c, v, timeperiod=14)
    df['adx'] = talib.ADX(h, l, c, timeperiod=14)
    df['pdi'] = talib.PLUS_DI(h, l, c, timeperiod=14)
    df['ndi'] = talib.MINUS_DI(h, l, c, timeperiod=14)
    bb_up, bb_mid, bb_low = talib.BBANDS(c, timeperiod=20, nbdevup=2, nbdevdn=2)
    df['bb_top'] = bb_up
    df['bb_mid'] = bb_mid
    df['bb_bot'] = bb_low
    df['bb_mid_dir'] = np.where(df['bb_mid'].diff() > 0, 'up', 
                                 np.where(df['bb_mid'].diff() < 0, 'down', 'flat'))
    return df

def run_backtest(df, params):
    """
    逐根 K 线回测
    params: {'adx':20, 'di_diff':5, 'bb_exit':8, 'time_stop':12}
    """
    adx_gate = params['adx']
    di_diff_gate = params['di_diff']
    bb_exit_offset = params['bb_exit']
    time_stop_bars = params.get('time_stop', 48)
    rsi_os = params.get('rsi_os', 30)     # 超卖阈值
    rsi_ob = params.get('rsi_ob', 70)     # 超买阈值
    mfi_os = params.get('mfi_os', 25)     # MFI 超卖阈值
    mfi_ob = params.get('mfi_ob', 75)     # MFI 超买阈值
    
    trades = []
    in_position = False
    direction = None  # 'LONG' or 'SHORT'
    entry_price = 0.0
    entry_time = None
    entry_bar = -1
    rsi_extreme = False
    mfi_extreme = False
    first_extreme_bar = -1
    pending_order = None  # 'buy' or 'sell'
    
    for i in range(len(df)):
        row = df.iloc[i]
        open_p = row['open']
        close_p = row['close']
        rsi = row['rsi']
        mfi = row['mfi']
        adx = row['adx']
        pdi = row['pdi']
        ndi = row['ndi']
        bb_top = row['bb_top']
        bb_bot = row['bb_bot']
        bb_mid_dir = row['bb_mid_dir']
        
        if np.isnan(rsi) or np.isnan(mfi) or np.isnan(adx):
            continue
        
        # 执行挂单（+1 bar 延迟入场）
        if pending_order == 'buy' and not in_position:
            entry_price = open_p
            entry_time = row['time']
            direction = 'LONG'
            in_position = True
            entry_bar = i
            rsi_extreme = False
            mfi_extreme = False
            first_extreme_bar = -1
            pending_order = None
            continue
        elif pending_order == 'sell' and not in_position:
            entry_price = open_p
            entry_time = row['time']
            direction = 'SHORT'
            in_position = True
            entry_bar = i
            rsi_extreme = False
            mfi_extreme = False
            first_extreme_bar = -1
            pending_order = None
            continue
        
        # ── 持仓中：检查出场 ──
        if in_position:
            bars_held = i - entry_bar
            
            if direction == 'LONG':
                # 跟踪 RSI/MFI 极限
                if not rsi_extreme and rsi >= rsi_ob:
                    rsi_extreme = True
                    if first_extreme_bar == -1:
                        first_extreme_bar = i
                if not mfi_extreme and mfi >= mfi_ob:
                    mfi_extreme = True
                    if first_extreme_bar == -1:
                        first_extreme_bar = i
                
                both_extreme = rsi_extreme and mfi_extreme
                
                # 时间止损：一个到了极限，另一个 N 根内没到
                if not both_extreme and first_extreme_bar != -1 and (i - first_extreme_bar) >= time_stop_bars:
                    exit_price = close_p
                    pnl = exit_price - entry_price
                    trades.append({'entry_time': entry_time, 'exit_time': row['time'], 'direction': 'LONG', 'entry_price': round(entry_price,2),
                                   'exit_price': round(exit_price,2), 'bars': bars_held, 'pnl': round(pnl,2),
                                   'reason': 'time_stop'})
                    in_position = False; direction = None; continue
                
                # 吃鱼出场：两个都到了极限，然后任一离开 + 价格回到 BB 上轨下方
                if both_extreme:
                    if (rsi < rsi_ob or mfi < mfi_ob) and close_p < bb_top - bb_exit_offset:
                        exit_price = close_p
                        pnl = exit_price - entry_price
                        trades.append({'entry_time': entry_time, 'exit_time': row['time'], 'direction': 'LONG', 'entry_price': round(entry_price,2),
                                       'exit_price': round(exit_price,2), 'bars': bars_held, 'pnl': round(pnl,2),
                                       'reason': 'fish_exit'})
                        in_position = False; direction = None; continue
            
            else:  # SHORT
                if not rsi_extreme and rsi <= rsi_os:
                    rsi_extreme = True
                    if first_extreme_bar == -1:
                        first_extreme_bar = i
                if not mfi_extreme and mfi <= mfi_os:
                    mfi_extreme = True
                    if first_extreme_bar == -1:
                        first_extreme_bar = i
                
                both_extreme = rsi_extreme and mfi_extreme
                
                # 时间止损：一个到了极限，另一个 N 根内没到
                if not both_extreme and first_extreme_bar != -1 and (i - first_extreme_bar) >= time_stop_bars:
                    exit_price = close_p
                    pnl = entry_price - exit_price
                    trades.append({'entry_time': entry_time, 'exit_time': row['time'], 'direction': 'SHORT', 'entry_price': round(entry_price,2),
                                   'exit_price': round(exit_price,2), 'bars': bars_held, 'pnl': round(pnl,2),
                                   'reason': 'time_stop'})
                    in_position = False; direction = None; continue
                
                if both_extreme:
                    if (rsi > rsi_os or mfi > mfi_os) and close_p > bb_bot + bb_exit_offset:
                        exit_price = close_p
                        pnl = entry_price - exit_price
                        trades.append({'entry_time': entry_time, 'exit_time': row['time'], 'direction': 'SHORT', 'entry_price': round(entry_price,2),
                                       'exit_price': round(exit_price,2), 'bars': bars_held, 'pnl': round(pnl,2),
                                       'reason': 'fish_exit'})
                        in_position = False; direction = None; continue
            
            continue  # 持仓中不做入场检查
        
        # ── 空仓：检查入场 ──
        if adx <= adx_gate:
            continue
        if abs(pdi - ndi) <= di_diff_gate:
            continue
        
        if ndi > pdi:  # -DI 大 → 空头主导 → 超卖 BUY
            if rsi < rsi_os and mfi < mfi_os and close_p <= bb_bot + 5 and bb_mid_dir == 'down':
                pending_order = 'buy'
        else:  # +DI 大 → 多头主导 → 超买 SELL
            if rsi > rsi_ob and mfi > mfi_ob and close_p >= bb_top - 5 and bb_mid_dir == 'up':
                pending_order = 'sell'
    
    return trades


def main():
    timeframes = ['M15', 'M30']
    param_grid = []
    for adx in [20, 22, 25]:
        for di_diff in [5, 10]:
            for bb_exit in [5, 8, 10]:
                param_grid.append({'adx': adx, 'di_diff': di_diff, 'bb_exit': bb_exit, 'time_stop': 48})
    
    # 样式
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='2F5496')
    green_fill = PatternFill('solid', fgColor='E2EFDA')
    red_fill = PatternFill('solid', fgColor='FCE4EC')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'), bottom=Side(style='thin'))
    
    wb = openpyxl.Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    for tf in timeframes:
        print(f'加载 {tf} 数据...')
        df = load_data(tf)
        print(f'  {len(df)} 根K线, 计算指标...')
        df = calc_indicators(df)
        print(f'  指标计算完成, 运行 {len(param_grid)} 组合...')
        
        all_results = []
        for i, p in enumerate(param_grid):
            combo = f"ADX{p['adx']}_DI{p['di_diff']}_BB{p['bb_exit']}_TS{p['time_stop']}"
            trades = run_backtest(df, p)
            
            total_trades = len(trades)
            wins = sum(1 for t in trades if t['pnl'] > 0)
            losses = sum(1 for t in trades if t['pnl'] <= 0)
            total_pnl = sum(t['pnl'] for t in trades)
            winrate = wins / total_trades * 100 if total_trades else 0
            
            all_results.append({
                'combo': combo, 'adx': p['adx'], 'di_diff': p['di_diff'],
                'bb_exit': p['bb_exit'], 'time_stop': p['time_stop'],
                'return': round(total_pnl, 2), 'trades': total_trades,
                'wins': wins, 'losses': losses, 'winrate': round(winrate, 1),
                'trades_list': trades
            })
            print(f'  [{i+1}/{len(param_grid)}] {combo} → {total_trades}笔 | 胜率{winrate:.0f}% | PnL:${total_pnl:+.2f}')
        
        all_results.sort(key=lambda r: r['return'], reverse=True)
        
        # ── 汇总 Sheet ──
        ws = wb.create_sheet(title=tf)
        ws.merge_cells('A1:J1')
        ws['A1'] = f'RSI-BB Trend 价格回归 · [{tf}]'
        ws['A1'].font = Font(bold=True, size=14, color='2F5496')
        ws.merge_cells('A2:J2')
        ws['A2'] = '入场: 门禁+3层筛子 | 出场: 吃鱼+时间止损 | 无硬止损 | TA-Lib 纯 Python 回测'
        ws['A2'].font = Font(italic=True, color='666666')
        
        headers = ['组合', 'ADX', 'DI', 'BB偏移', 'TS', '净PnL($)', '交易', '胜率(%)', '胜笔', '亏笔']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font = header_font; cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center'); cell.border = thin_border
        
        for i, r in enumerate(all_results, 5):
            vals = [r['combo'], r['adx'], r['di_diff'], r['bb_exit'], r['time_stop'],
                    r['return'], r['trades'], r['winrate'], r['wins'], r['losses']]
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=i, column=col, value=v)
                cell.border = thin_border; cell.alignment = Alignment(horizontal='center')
                if col == 6:
                    if v > 0: cell.fill = green_fill; cell.font = Font(color='006100')
                    elif v < 0: cell.fill = red_fill; cell.font = Font(color='9C0006')
        
        widths = [32, 8, 8, 10, 6, 14, 8, 10, 8, 8]
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w
        for col in range(1, 11):
            ws.cell(row=5, column=col).font = Font(bold=True, color='B8860B')
        
        # 分组统计
        gs = len(all_results) + 7
        ws.merge_cells(f'A{gs}:G{gs}')
        ws.cell(row=gs, column=1, value='分组统计').font = Font(bold=True, size=12, color='2F5496')
        sr = gs + 1
        for dim_name, dim_key in [('ADX', 'adx'), ('DI差值', 'di_diff'), ('BB偏移', 'bb_exit')]:
            groups = {}
            for r in all_results:
                val = r[dim_key]
                if val not in groups:
                    groups[val] = {'pnls': [], 'trades': 0, 'wins': 0}
                groups[val]['pnls'].append(r['return'])
                groups[val]['trades'] += r['trades']
                groups[val]['wins'] += r['wins']
            for h, c in [('参数',1), ('组合数',2), ('平均PnL($)',3), ('总交易',4), ('总胜笔',5), ('平均胜率(%)',6)]:
                cell = ws.cell(row=sr, column=c, value=h); cell.font = Font(bold=True); cell.border = thin_border
            sr += 1
            for val in sorted(groups.keys()):
                g = groups[val]; avg_pnl = sum(g['pnls']) / len(g['pnls'])
                avg_wr = g['wins'] / g['trades'] * 100 if g['trades'] else 0
                for c, v in [(1,val),(2,len(g['pnls'])),(3,round(avg_pnl,2)),(4,g['trades']),(5,g['wins']),(6,round(avg_wr,1))]:
                    cell = ws.cell(row=sr, column=c, value=v); cell.border = thin_border; cell.alignment = Alignment(horizontal='center')
                    if c == 3:
                        if v > 0: cell.fill = green_fill; cell.font = Font(color='006100')
                        elif v < 0: cell.fill = red_fill; cell.font = Font(color='9C0006')
                sr += 1
            sr += 1
        
        # ── 明细 Sheet ──
        ds = wb.create_sheet(title=f'{tf}_明细')
        dheaders = ['组合', '方向', '入场时间', '入场价', '出场时间', '出场价', '持仓K线', 'PnL($)', '出场原因']
        for col, h in enumerate(dheaders, 1):
            cell = ds.cell(row=1, column=col, value=h)
            cell.font = header_font; cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center'); cell.border = thin_border
        
        dr = 2
        for r in all_results:
            for t in r['trades_list']:
                vals = [r['combo'], t['direction'], t['entry_time'].strftime('%Y-%m-%d %H:%M'),
                        t['entry_price'], t['exit_time'].strftime('%Y-%m-%d %H:%M'), t['exit_price'],
                        t['bars'], t['pnl'], t['reason']]
                for col, v in enumerate(vals, 1):
                    cell = ds.cell(row=dr, column=col, value=v)
                    cell.border = thin_border; cell.alignment = Alignment(horizontal='center')
                    if col == 8:
                        if v > 0: cell.fill = green_fill
                        elif v < 0: cell.fill = red_fill
                dr += 1
        
        dwidths = [28, 10, 18, 12, 18, 12, 12, 12, 12]
        for col, w in enumerate(dwidths, 1):
            ds.column_dimensions[get_column_letter(col)].width = w
        ds.auto_filter.ref = f'A1:I{dr-1}'
    
    out_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                            'backtest', 'rsi_bb_trend_results_ts48.xlsx')
    wb.save(out_path)
    print(f'\n✅ Excel: {out_path}')


if __name__ == '__main__':
    main()