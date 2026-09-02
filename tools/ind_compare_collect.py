#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
指标比对采集脚本（独立直连 MT4，8 小时连续采集）
每 1 分钟采集一次 M15/M30/H1 三个周期：
  - EA 原始值（F043 命令，与 MT4 图表完全一致）
  - TA-Lib 独立计算值（从 DB OHLCV 重新计算）
实时写入 Excel，采集结束后生成汇总分析 Sheet。
用法: python tools/ind_compare_collect.py
"""
import sys, os, time, json, math, logging, signal as sig_mod
from datetime import datetime, timezone, timedelta
from pathlib import Path

# 项目根目录加入 path
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import numpy as np
import talib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from core.freemt4_bridge import FreeMT4Bridge

# ── 配置 ──────────────────────────────────────────────
DURATION_HOURS = 8
INTERVAL_SEC = 60
TIMEFRAMES = ["M15", "M30", "H1"]
SYMBOL = "XAUUSD"
# ohlcv.timestamp 是 MT4 服务器墙钟（UTC+3）直接当 epoch 存
# 即 stored_ts = real_epoch + 3*3600
MT4_TZ_OFFSET = 3 * 3600
LOCAL_TZ = timezone(timedelta(hours=8))
OUTPUT_DIR = ROOT / "tools" / "indicator_compare_output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
EXCEL_PATH = OUTPUT_DIR / "indicator_compare.xlsx"
LOG_PATH = OUTPUT_DIR / "collect.log"

# ── 日志 ──────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.FileHandler(LOG_PATH, encoding="utf-8"), logging.StreamHandler()],
)
log = logging.getLogger("collect")

# ── EA 提供的可比对指标（F043 返回的标量字段）─────────
# 每项: (excel列名, EA字典key, 取值函数/None, 对应talib计算函数名)
SCALAR_INDICATORS = [
    ("rsi",        "rsi",        lambda ea: ea.get("rsi"),        "rsi"),
    ("rsi_5",      "rsi_5",     lambda ea: ea.get("rsi_5"),      "rsi_5"),
    ("rsi_10",     "rsi_10",    lambda ea: ea.get("rsi_10"),     "rsi_10"),
    ("mfi",        "mfi",       lambda ea: ea.get("mfi"),        "mfi"),
    ("ema_9",      "ema_9",     lambda ea: ea.get("ema_9"),      "ema_9"),
    ("ema_21",     "ema_21",    lambda ea: ea.get("ema_21"),     "ema_21"),
    ("sma_14",     "sma_14",    lambda ea: ea.get("sma_14"),     "sma_14"),
    ("sma_20",     "sma_20",    lambda ea: ea.get("sma_20"),     "sma_20"),
    ("sma_50",     "sma_50",    lambda ea: ea.get("sma_50"),     "sma_50"),
    ("atr",        "atr",       lambda ea: ea.get("atr"),        "atr"),
    ("atr_20",     "atr_20",    lambda ea: ea.get("atr_20"),     "atr_20"),
    ("adx",        "adx",       lambda ea: ea.get("adx"),        "adx"),
    ("pdi",        "pdi",       lambda ea: ea.get("pdi"),         "pdi"),
    ("ndi",        "ndi",       lambda ea: ea.get("ndi"),         "ndi"),
    ("volume_sma_20", "volume_sma_20", lambda ea: ea.get("volume_sma_20"), "volume_sma_20"),
]
# 子字段指标（BB / MACD / Stoch 各有子键）
SUB_INDICATORS = [
    ("bb_upper",   "bb",   lambda ea: ea.get("bb",{}).get("upper"),   "bb_upper"),
    ("bb_mid",     "bb",   lambda ea: ea.get("bb",{}).get("mid"),     "bb_mid"),
    ("bb_lower",   "bb",   lambda ea: ea.get("bb",{}).get("lower"),   "bb_lower"),
    ("macd_macd",  "macd", lambda ea: ea.get("macd",{}).get("macd"),  "macd_macd"),
    ("macd_signal","macd", lambda ea: ea.get("macd",{}).get("signal"),"macd_signal"),
    ("stoch_k",    "stoch_5_3_3", lambda ea: ea.get("stoch_5_3_3",{}).get("k"), "stoch_k"),
    ("stoch_d",    "stoch_5_3_3", lambda ea: ea.get("stoch_5_3_3",{}).get("d"), "stoch_d"),
]
ALL_INDICATORS = SCALAR_INDICATORS + SUB_INDICATORS

# ── TA-Lib 独立计算 ──────────────────────────────────
def talib_calculate(candles: list, key: str):
    """根据指标 key 用 talib 独立计算最新一根 K 线的值。
    candles: list[Candle]（Candle 有 .time/.open/.high/.low/.close/.volume），按时间升序
    返回 float 或 None
    """
    if not candles or len(candles) < 50:
        return None
    # EA 的 get_candles 返回降序（最新在前），talib 需要升序（旧→新）
    candles = list(reversed(candles))
    closes = np.array([c.close for c in candles], dtype=float)
    highs  = np.array([c.high  for c in candles], dtype=float)
    lows   = np.array([c.low   for c in candles], dtype=float)
    vols   = np.array([getattr(c, 'volume', 0) for c in candles], dtype=float)
    i = len(candles) - 1  # 最新一根（与 EA time 对应的当前 bar）
    try:
        if key == "rsi":
            v = talib.RSI(closes, timeperiod=14)[i]
        elif key == "rsi_5":
            v = talib.RSI(closes, timeperiod=5)[i]
        elif key == "rsi_10":
            v = talib.RSI(closes, timeperiod=10)[i]
        elif key == "mfi":
            v = talib.MFI(highs, lows, closes, vols, timeperiod=14)[i]
        elif key == "ema_9":
            v = talib.EMA(closes, timeperiod=9)[i]
        elif key == "ema_21":
            v = talib.EMA(closes, timeperiod=21)[i]
        elif key == "sma_14":
            v = talib.SMA(closes, timeperiod=14)[i]
        elif key == "sma_20":
            v = talib.SMA(closes, timeperiod=20)[i]
        elif key == "sma_50":
            v = talib.SMA(closes, timeperiod=50)[i]
        elif key == "atr":
            v = talib.ATR(highs, lows, closes, timeperiod=14)[i]
        elif key == "atr_20":
            v = talib.ATR(highs, lows, closes, timeperiod=20)[i]
        elif key == "adx":
            v = talib.ADX(highs, lows, closes, timeperiod=14)[i]
        elif key == "pdi":
            v = talib.PLUS_DI(highs, lows, closes, timeperiod=14)[i]
        elif key == "ndi":
            v = talib.MINUS_DI(highs, lows, closes, timeperiod=14)[i]
        elif key == "volume_sma_20":
            v = talib.SMA(vols, timeperiod=20)[i]
        elif key == "bb_upper":
            up, mid, lo = talib.BBANDS(closes, timeperiod=20, nbdevup=2, nbdevdn=2)
            v = up[i]
        elif key == "bb_mid":
            up, mid, lo = talib.BBANDS(closes, timeperiod=20, nbdevup=2, nbdevdn=2)
            v = mid[i]
        elif key == "bb_lower":
            up, mid, lo = talib.BBANDS(closes, timeperiod=20, nbdevup=2, nbdevdn=2)
            v = lo[i]
        elif key == "macd_macd":
            m, s, _ = talib.MACD(closes, fastperiod=12, slowperiod=26, signalperiod=9)
            v = m[i]
        elif key == "macd_signal":
            m, s, _ = talib.MACD(closes, fastperiod=12, slowperiod=26, signalperiod=9)
            v = s[i]
        elif key == "stoch_k":
            sk, sd = talib.STOCH(highs, lows, closes, fastk_period=5, slowk_period=3, slowd_period=3)
            v = sk[i]
        elif key == "stoch_d":
            sk, sd = talib.STOCH(highs, lows, closes, fastk_period=5, slowk_period=3, slowd_period=3)
            v = sd[i]
        else:
            return None
        return float(v) if v == v else None  # NaN 检查
    except Exception:
        return None

# ── Excel 初始化 ────────────────────────────────────
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
WARN_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

def build_header(indicators):
    h = ["采集时间", "EA_K线时间", "tf", "close"]
    for name, _, _, _ in indicators:
        h += [f"{name}_EA", f"{name}_talib", f"{name}_diff", f"{name}_match"]
    return h

def init_workbook():
    wb = Workbook()
    wb.remove(wb.active)
    for tf in TIMEFRAMES:
        ws = wb.create_sheet(tf)
        header = build_header(ALL_INDICATORS)
        ws.append(header)
        for col in range(1, len(header) + 1):
            c = ws.cell(row=1, column=col)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = Alignment(horizontal="center")
    wb.save(EXCEL_PATH)
    return wb

# ── 采集主循环 ──────────────────────────────────────
def main():
    log.info("=" * 60)
    log.info("指标比对采集启动 | 时长 %dh | 间隔 %ds | 周期 %s", DURATION_HOURS, INTERVAL_SEC, TIMEFRAMES)
    log.info("输出: %s", EXCEL_PATH)
    log.info("=" * 60)

    wb = init_workbook()
    total_rounds = int(DURATION_HOURS * 3600 / INTERVAL_SEC)
    round_idx = 0

    bridge = FreeMT4Bridge()
    connected = False

    def ensure_connect():
        nonlocal bridge, connected
        if connected and bridge._connected:
            return True
        log.info("连接 MT4 EA ...")
        if bridge.connect():
            connected = True
            log.info("MT4 EA 已连接")
            return True
        log.error("MT4 EA 连接失败")
        return False

    # 优雅退出
    def on_sig(signum, frame):
        log.info("收到退出信号 %s，保存并退出", signum)
        wb.save(EXCEL_PATH)
        try: bridge.disconnect()
        except: pass
        sys.exit(0)
    sig_mod.signal(sig_mod.SIGINT, on_sig)
    if hasattr(sig_mod, 'SIGBREAK'):
        sig_mod.signal(sig_mod.SIGBREAK, on_sig)

    try:
        while round_idx < total_rounds:
            round_idx += 1
            ts_now = datetime.now(LOCAL_TZ)
            log.info("━━ 第 %d/%d 轮 %s ━━", round_idx, total_rounds, ts_now.strftime("%Y-%m-%d %H:%M:%S"))

            for tf in TIMEFRAMES:
                try:
                    if not ensure_connect():
                        log.warning("跳过 %s (连接失败)", tf)
                        continue

                    # 1. EA 原始值
                    ea = bridge.get_indicators(SYMBOL, tf)
                    if not ea:
                        log.warning("F043 %s 返回空", tf)
                        continue
                    ea_time = ea.get("time")
                    ea_close = ea.get("close")

                    # 2. TA-Lib 独立计算（K线也从 EA 取，与指标完全同源同时戳）
                    candles = bridge.get_candles(SYMBOL, tf, count=300, offset=0)

                    # 3. 构建一行
                    row = [ts_now.strftime("%Y-%m-%d %H:%M:%S"),
                           str(ea_time) if ea_time else "", tf, ea_close]
                    ws = wb[tf]
                    for name, _ea_key, extract_fn, talib_key in ALL_INDICATORS:
                        ea_val = extract_fn(ea)
                        tl_val = talib_calculate(candles, talib_key) if candles else None
                        diff = None
                        match = ""
                        if ea_val is not None and tl_val is not None:
                            diff = round(ea_val - tl_val, 6)
                            match = "OK" if abs(diff) < 0.01 else "DIFF"
                        row.append(ea_val)
                        row.append(tl_val)
                        row.append(diff)
                        row.append(match)
                    ws.append(row)
                    # 标红超限行
                    r = ws.max_row
                    diff_cols = [4 + 3 + i*4 for i in range(len(ALL_INDICATORS))]  # diff 列偏移
                    for dc in diff_cols:
                        cell = ws.cell(row=r, column=dc)
                        if cell.value is not None and abs(float(cell.value)) >= 0.01:
                            cell.fill = WARN_FILL
                    wb.save(EXCEL_PATH)

                    # 简要日志
                    k_ea = ea.get("stoch_5_3_3",{}).get("k")
                    k_tl = talib_calculate(candles, "stoch_k") if candles else None
                    log.info("%s EA_K=%s tl_K=%s close=%s candles=%d", tf, k_ea, k_tl, ea_close, len(candles))

                except Exception as e:
                    log.exception("%s 采集异常: %s", tf, e)
                    # 连接可能断了，重置
                    connected = False
                    try: bridge.disconnect()
                    except: pass

            # 本轮结束，等待下一轮
            if round_idx < total_rounds:
                time.sleep(INTERVAL_SEC)

    finally:
        wb.save(EXCEL_PATH)
        log.info("采集结束，Excel 已保存: %s", EXCEL_PATH)
        log.info("开始生成汇总...")
        try:
            generate_summary()
        except Exception as e:
            log.exception("汇总生成失败: %s", e)
        try: bridge.disconnect()
        except: pass

# ── 汇总分析 ────────────────────────────────────────
def generate_summary():
    from openpyxl import load_workbook
    wb = load_workbook(EXCEL_PATH)
    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    ws = wb.create_sheet("Summary", 0)
    ws.append(["指标比对汇总分析", "", "", "", "", ""])
    ws.append([f"生成时间: {datetime.now(LOCAL_TZ).strftime('%Y-%m-%d %H:%M:%S')}"])
    ws.append([])
    ws.append(["周期", "指标", "样本数", "匹配数", "超限数",
               "匹配率%", "平均误差", "最大误差", "RMSE", "超限阈值0.01"])
    for col in range(1, 11):
        ws.cell(row=4, column=col).fill = HEADER_FILL
        ws.cell(row=4, column=col).font = HEADER_FONT

    r = 5
    for tf in TIMEFRAMES:
        if tf not in wb.sheetnames:
            continue
        src = wb[tf]
        rows = list(src.iter_rows(min_row=2, values_only=True))
        if not rows:
            continue
        n_cols = len(rows[0])
        # 每个指标占 4 列: EA, talib, diff, match
        for idx, (name, _, _, _) in enumerate(ALL_INDICATORS):
            base = 4 + idx * 4  # 0-based: EA列
            diffs = []
            matches = 0
            total = 0
            for row in rows:
                if len(row) <= base + 2:
                    continue
                d = row[base + 2]
                m = row[base + 3]
                if d is not None and m:
                    total += 1
                    diffs.append(float(d))
                    if m == "OK":
                        matches += 1
            if total == 0:
                continue
            overs = total - matches
            mean_e = sum(diffs) / len(diffs)
            max_e = max(abs(x) for x in diffs)
            rmse = math.sqrt(sum(x*x for x in diffs) / len(diffs))
            ws.append([tf, name, total, matches, overs,
                       round(matches/total*100, 2),
                       round(mean_e, 6), round(max_e, 6), round(rmse, 6), 0.01])
            r += 1

    # 列宽自适应
    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["J"].width = 16

    wb.save(EXCEL_PATH)
    log.info("汇总 Sheet 已生成: %s", EXCEL_PATH)

if __name__ == "__main__":
    main()
