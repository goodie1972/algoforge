#!/c/Python314/python
"""
XAUUSD 手动策略模拟系统 v2.0 - Excel 完整报表版
跑在实盘信号上，不执行操作。每15分钟检查一次。
"""

import json, os, sys, math, time, csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers

# ── 配置 ────────────────────────────────────────────────
INITIAL_CAPITAL = 10000.0      # 模拟初始资金 USD
FIXED_LOT = 0.1                # 固定每笔 0.1手
ENGINE_URL = "http://localhost:1783"
DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
os.makedirs(DATA_DIR, exist_ok=True)

STATE_FILE = os.path.join(DATA_DIR, "sim_state.json")
SCORE_LOG_FILE = os.path.join(DATA_DIR, "factor_scores.csv")
TRADE_LOG_FILE = os.path.join(DATA_DIR, "trade_log.csv")
EXCEL_REPORT = os.path.join(os.path.dirname(__file__), "模拟交易报表.xlsx")


# ── 指标计算 ────────────────────────────────────────────

def ema(data, period):
    """计算EMA"""
    if len(data) < period:
        return data
    k = 2.0 / (period + 1)
    ema_val = sum(data[:period]) / period
    result = [ema_val] * (period - 1)  # pad
    result.append(ema_val)
    for price in data[period:]:
        ema_val = price * k + ema_val * (1 - k)
        result.append(round(ema_val, 2))
    return result

def rsi(closes, period=14):
    """计算RSI"""
    if len(closes) < period + 1:
        return [50] * len(closes)
    gains, losses = [], []
    for i in range(1, len(closes)):
        diff = closes[i] - closes[i-1]
        gains.append(max(diff, 0))
        losses.append(max(-diff, 0))
    
    result = [50] * period  # pad
    avg_gain = sum(gains[:period]) / period
    avg_loss = sum(losses[:period]) / period
    for i in range(period, len(gains)):
        if avg_loss == 0:
            result.append(100.0)
        else:
            rs = avg_gain / avg_loss
            result.append(round(100 - 100 / (1 + rs), 1))
        avg_gain = (avg_gain * (period - 1) + gains[i]) / period
        avg_loss = (avg_loss * (period - 1) + losses[i]) / period
    return result

def sma(data, period):
    if len(data) < period:
        return [sum(data)/len(data)] * len(data)
    result = [0] * len(data)
    cum = sum(data[:period])
    result[period-1] = round(cum / period, 2)
    for i in range(period, len(data)):
        cum = cum - data[i-period] + data[i]
        result[i] = round(cum / period, 2)
    for i in range(period-1):
        result[i] = result[period-1]
    return result

def bollinger(closes, period=20, std_dev=2):
    """Returns (middle, upper, lower, bandwidth) arrays"""
    ma = sma(closes, period)
    upper, lower, bandwidth = [], [], []
    for i in range(len(closes)):
        if i < period-1:
            upper.append(round(ma[i], 2))
            lower.append(round(ma[i], 2))
            bandwidth.append(0)
            continue
        window = closes[i-period+1:i+1]
        sd = (sum((x - ma[i])**2 for x in window) / period) ** 0.5
        upper.append(round(ma[i] + std_dev * sd, 2))
        lower.append(round(ma[i] - std_dev * sd, 2))
        bw = (upper[-1] - lower[-1]) / ma[i] * 100 if ma[i] > 0 else 0
        bandwidth.append(round(bw, 2))
    return ma, upper, lower, bandwidth

def atr(candles, period=14):
    """计算ATR"""
    if len(candles) < period + 1:
        return 0
    trs = []
    for i in range(1, len(candles)):
        high, low = candles[i]['high'], candles[i]['low']
        prev_close = candles[i-1]['close']
        tr = max(high-low, abs(high-prev_close), abs(low-prev_close))
        trs.append(tr)
    return sum(trs[-period:]) / period if trs else 0


# ── 因子评分系统 ─────────────────────────────────────────

def compute_factors(m30, h1, curr_price):
    """计算所有策略的因子, 返回完整的因子字典"""
    if len(m30) < 50 or len(h1) < 30:
        return None
    
    closes_m30 = [b['close'] for b in m30]
    closes_h1 = [b['close'] for b in h1]
    
    now_ts = m30[-1]['time']
    now_str = datetime.fromtimestamp(now_ts).strftime("%Y-%m-%d %H:%M:%S")
    
    # ─── EMA (H1) ───
    ema9_h1 = ema(closes_h1, 9)
    ema21_h1 = ema(closes_h1, 21)
    ema9_curr = ema9_h1[-1]
    ema21_curr = ema21_h1[-1]
    ema_prev9, ema_prev21 = ema9_h1[-2], ema21_h1[-2]
    
    # ─── RSI (M30) ───
    rsi_vals = rsi(closes_m30, 14)
    rsi_curr = rsi_vals[-1]
    
    # ─── BB (M30) ───
    bb_mid, bb_upper, bb_lower, bb_bw = bollinger(closes_m30, 20, 2)
    bb_mid_curr = bb_mid[-1]
    bb_up_curr = bb_upper[-1]
    bb_lo_curr = bb_lower[-1]
    bb_bw_curr = bb_bw[-1]
    bb_bw_avg = sum(bb_bw[-20:]) / 20 if len(bb_bw) >= 20 else bb_bw_curr
    bb_squeeze = bb_bw_curr < bb_bw_avg * 0.7 if bb_bw_avg > 0 else False
    
    # ─── ADX (H1) ───
    adx_vals, pdi_vals, ndi_vals = adx(h1, 14)
    adx_curr = adx_vals[-1]
    pdi_curr = pdi_vals[-1]
    ndi_curr = ndi_vals[-1]
    
    # 当前价格
    bid = curr_price.get('bid', m30[-1]['close'])
    
    # ═══ 各策略因子评分 ═══
    factors = {}
    
    # ── 策略1: EMA CROSS ──
    ema_cross_score = {}
    ema_cross_score['EMA9'] = ema9_curr
    ema_cross_score['EMA21'] = ema21_curr
    ema_cross_score['价差'] = round(ema9_curr - ema21_curr, 2)
    if ema_prev9 <= ema_prev21 and ema9_curr > ema21_curr:
        ema_cross_score['信号'] = 'BUY'
        ema_cross_score['评分'] = 3
        ema_cross_score['说明'] = f'金叉: EMA9({ema9_curr})上穿EMA21({ema21_curr})'
    elif ema_prev9 >= ema_prev21 and ema9_curr < ema21_curr:
        ema_cross_score['信号'] = 'SELL'
        ema_cross_score['评分'] = 3
        ema_cross_score['说明'] = f'死叉: EMA9({ema9_curr})下穿EMA21({ema21_curr})'
    else:
        ema_cross_score['信号'] = '无'
        ema_cross_score['评分'] = 0
        ema_cross_score['说明'] = f'EMA9({ema9_curr})运行中，暂未交叉'
        if ema9_curr > ema21_curr:
            ema_cross_score['说明'] += '(多头排列)'
        else:
            ema_cross_score['说明'] += '(空头排列)'
    factors['EMA_CROSS'] = ema_cross_score
    
    # ── 策略2: RSI REV ──
    rsi_rev_score = {}
    trend_up = ema9_h1[-1] > ema9_h1[-4] if len(ema9_h1) >= 4 else False
    trend_down = ema9_h1[-1] < ema9_h1[-4] if len(ema9_h1) >= 4 else False
    rsi_rev_score['RSI'] = rsi_curr
    rsi_rev_score['H1趋势'] = '↑' if trend_up else ('↓' if trend_down else '→')
    rsi_rev_score['趋势方向'] = '向上' if trend_up else ('向下' if trend_down else '横盘')
    rsi_rev_score['H1_EMA9斜率'] = round(ema9_h1[-1] - ema9_h1[-4], 2) if len(ema9_h1) >= 4 else 0
    
    score = 0
    if rsi_curr < 30 and trend_up:
        rsi_rev_score['信号'] = 'BUY'
        score = 3
        rsi_rev_score['说明'] = f'RSI超卖({rsi_curr})+趋势向上, 做多信号'
    elif rsi_curr > 70 and trend_down:
        rsi_rev_score['信号'] = 'SELL'
        score = 3
        rsi_rev_score['说明'] = f'RSI超买({rsi_curr})+趋势向下, 做空信号'
    elif 40 <= rsi_curr <= 60:
        rsi_rev_score['信号'] = 'EXIT'
        score = 0
        rsi_rev_score['说明'] = f'RSI回到中性区({rsi_curr}), 平仓'
    else:
        if rsi_curr < 30:
            rsi_rev_score['信号'] = '无'
            rsi_rev_score['说明'] = f'RSI超卖但趋势不支持'
        elif rsi_curr > 70:
            rsi_rev_score['信号'] = '无'
            rsi_rev_score['说明'] = f'RSI超买但趋势不支持'
        else:
            rsi_rev_score['信号'] = '无'
            rsi_rev_score['说明'] = f'RSI中性区({rsi_curr})等待'
    rsi_rev_score['评分'] = score
    factors['RSI_REV'] = rsi_rev_score
    
    # ── 策略3: BB SQUEEZE ──
    bb_squeeze_score = {}
    bb_squeeze_score['BB_上轨'] = bb_up_curr
    bb_squeeze_score['BB_中轨'] = bb_mid_curr
    bb_squeeze_score['BB_下轨'] = bb_lo_curr
    bb_squeeze_score['BB带宽(%)'] = bb_bw_curr
    bb_squeeze_score['BB带宽均值(%)'] = round(bb_bw_avg, 2)
    bb_squeeze_score['挤仓状态'] = '是' if bb_squeeze else '否'
    
    score = 0
    if bb_squeeze and bid > bb_up_curr:
        bb_squeeze_score['信号'] = 'BUY'
        score = 3
        bb_squeeze_score['说明'] = f'挤仓突破上轨: 价格{bid} > BB上{bb_up_curr}'
    elif bb_squeeze and bid < bb_lo_curr:
        bb_squeeze_score['信号'] = 'SELL'
        score = 3
        bb_squeeze_score['说明'] = f'挤仓突破下轨: 价格{bid} < BB下{bb_lo_curr}'
    elif not bb_squeeze and bb_lo_curr <= bid <= bb_up_curr:
        bb_squeeze_score['信号'] = 'EXIT'
        score = 0
        bb_squeeze_score['说明'] = f'价格回到BB内部, 平仓信号'
    else:
        bb_squeeze_score['信号'] = '无'
        if bb_squeeze:
            bb_squeeze_score['说明'] = '挤仓中但未突破'
        else:
            bb_squeeze_score['说明'] = '未挤仓'
    bb_squeeze_score['评分'] = score
    factors['BB_SQUEEZE'] = bb_squeeze_score
    
    # ── 策略4: ADX RIDER ──
    adx_rider_score = {}
    adx_rider_score['ADX'] = round(adx_curr, 1)
    adx_rider_score['DI+'] = round(pdi_curr, 1)
    adx_rider_score['DI-'] = round(ndi_curr, 1)
    adx_rider_score['DI差'] = round(pdi_curr - ndi_curr, 1)
    
    score = 0
    if adx_curr > 25 and pdi_curr > ndi_curr:
        adx_rider_score['信号'] = 'BUY'
        score = 3
        adx_rider_score['说明'] = f'强趋势(ADX={adx_curr:.1f}>25), DI+={pdi_curr:.1f}>DI-={ndi_curr:.1f}, 做多'
    elif adx_curr > 25 and ndi_curr > pdi_curr:
        adx_rider_score['信号'] = 'SELL'
        score = 3
        adx_rider_score['说明'] = f'强趋势(ADX={adx_curr:.1f}>25), DI-={ndi_curr:.1f}>DI+={pdi_curr:.1f}, 做空'
    elif adx_curr < 20:
        adx_rider_score['信号'] = 'EXIT'
        score = 0
        adx_rider_score['说明'] = f'ADX={adx_curr:.1f}<20, 趋势消失, 平仓'
    else:
        adx_rider_score['信号'] = '无'
        adx_rider_score['说明'] = f'ADX={adx_curr:.1f}介于20-25之间, 等待趋势确认'
    adx_rider_score['评分'] = score
    factors['ADX_RIDER'] = adx_rider_score
    
    return {
        'time': now_str,
        'timestamp': now_ts,
        'price': round(bid, 2),
        'factors': factors
    }


# ── ADX 计算（单独抽出来给factor用） ──
def adx(candles, period=14):
    if len(candles) < period + 1:
        return [0]*len(candles), [0]*len(candles), [0]*len(candles)
    trs, plus_dms, minus_dms = [], [], []
    for i in range(1, len(candles)):
        high, low = candles[i]['high'], candles[i]['low']
        prev_close = candles[i-1]['close']
        tr = max(high-low, abs(high-prev_close), abs(low-prev_close))
        trs.append(tr)
        up_move = high - candles[i-1]['high']
        down_move = candles[i-1]['low'] - low
        plus_dms.append(max(up_move, 0) if up_move > down_move else 0)
        minus_dms.append(max(down_move, 0) if down_move > up_move else 0)
    
    def smooth(values, n):
        result = [0] * len(values)
        if len(values) < n:
            return result
        k = 1.0 / n
        smoothed = sum(values[:n]) / n
        result[n-1] = smoothed
        for i in range(n, len(values)):
            smoothed = (smoothed * (n - 1) + values[i]) / n
            result[i] = smoothed
        return result
    
    tr_ema = smooth(trs, period)
    pdi_ema = smooth(plus_dms, period)
    ndi_ema = smooth(minus_dms, period)
    
    # Pad arrays to match candle count
    pdi_pad = 1
    pad_len = len(candles) - len(pdi_ema)
    
    dx_vals = []
    for i in range(len(pdi_ema)):
        di_sum = pdi_ema[i] + ndi_ema[i]
        if di_sum == 0 or tr_ema[i] == 0:
            dx_vals.append(0)
        else:
            dx = abs(pdi_ema[i] - ndi_ema[i]) / di_sum * 100
            dx_vals.append(dx)
    
    adx_raw = smooth(dx_vals, period)
    
    # Convert DI to percentage-like values
    max_val = max(max(tr_ema), 1)
    pdi_pct = [v / max_val * 100 if max_val > 0 else 0 for v in pdi_ema]
    ndi_pct = [v / max_val * 100 if max_val > 0 else 0 for v in ndi_ema]
    
    # Pad to match candle count (pad with first value or 0)
    full_adx = [0] * pad_len + adx_raw
    full_pdi = [0] * pad_len + pdi_pct
    full_ndi = [0] * pad_len + ndi_pct
    
    return full_adx, full_pdi, full_ndi


# ── 模拟仓管理 ───────────────────────────────────────────

def load_state():
    if os.path.exists(STATE_FILE):
        with open(STATE_FILE) as f:
            return json.load(f)
    return {
        "capital": INITIAL_CAPITAL,
        "equity": INITIAL_CAPITAL,
        "positions": [],
        "trades": [],
        "total_trades": 0,
        "wins": 0,
        "losses": 0
    }

def save_state(state):
    with open(STATE_FILE, 'w') as f:
        json.dump(state, f, indent=2)

def log_factor_scores(factor_data):
    """追加因子评分到CSV"""
    file_exists = os.path.exists(SCORE_LOG_FILE)
    with open(SCORE_LOG_FILE, 'a', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        if not file_exists:
            headers = [
                '时间', '价格',
                'EMA_CROSS_信号', 'EMA_CROSS_评分', 'EMA_CROSS_说明', 'EMA9', 'EMA21', '价差',
                'RSI_REV_信号', 'RSI_REV_评分', 'RSI_REV_说明', 'RSI', 'H1趋势', 'H1_EMA9斜率',
                'BB_SQUEEZE_信号', 'BB_SQUEEZE_评分', 'BB_SQUEEZE_说明', 'BB上轨', 'BB中轨', 'BB下轨', 'BB带宽%', '挤仓',
                'ADX_RIDER_信号', 'ADX_RIDER_评分', 'ADX_RIDER_说明', 'ADX', 'DI+', 'DI-', 'DI差'
            ]
            writer.writerow(headers)
        
        f = factor_data['factors']
        row = [
            factor_data['time'], factor_data['price'],
            f['EMA_CROSS']['信号'], f['EMA_CROSS']['评分'], f['EMA_CROSS']['说明'],
            f['EMA_CROSS']['EMA9'], f['EMA_CROSS']['EMA21'], f['EMA_CROSS']['价差'],
            f['RSI_REV']['信号'], f['RSI_REV']['评分'], f['RSI_REV']['说明'],
            f['RSI_REV']['RSI'], f['RSI_REV']['趋势方向'], f['RSI_REV']['H1_EMA9斜率'],
            f['BB_SQUEEZE']['信号'], f['BB_SQUEEZE']['评分'], f['BB_SQUEEZE']['说明'],
            f['BB_SQUEEZE']['BB_上轨'], f['BB_SQUEEZE']['BB_中轨'], f['BB_SQUEEZE']['BB_下轨'],
            f['BB_SQUEEZE']['BB带宽(%)'], f['BB_SQUEEZE']['挤仓状态'],
            f['ADX_RIDER']['信号'], f['ADX_RIDER']['评分'], f['ADX_RIDER']['说明'],
            f['ADX_RIDER']['ADX'], f['ADX_RIDER']['DI+'], f['ADX_RIDER']['DI-'], f['ADX_RIDER']['DI差']
        ]
        writer.writerow(row)

def log_trade(trade_data):
    """追加交易记录到CSV"""
    file_exists = os.path.exists(TRADE_LOG_FILE)
    with open(TRADE_LOG_FILE, 'a', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        if not file_exists:
            headers = ['时间', '策略', '动作', '方向', '价格', '手数', '盈亏($)', '原因']
            writer.writerow(headers)
        writer.writerow([
            trade_data['time'],
            trade_data['strategy'],
            trade_data['action'],
            trade_data.get('direction', ''),
            trade_data['price'],
            trade_data.get('lots', FIXED_LOT),
            trade_data.get('pnl', 0),
            trade_data.get('reason', '')
        ])


def open_position(state, strategy, direction, price, time_now):
    pos = {
        "strategy": strategy,
        "direction": direction,
        "open_price": price,
        "lots": FIXED_LOT,
        "open_time": time_now
    }
    state['positions'].append(pos)
    log_trade({
        'time': time_now, 'strategy': strategy, 'action': '开仓',
        'direction': direction, 'price': price, 'lots': FIXED_LOT,
        'pnl': 0, 'reason': f'{strategy}信号触发'
    })
    return pos

def close_position(state, pos, close_price, reason, time_now):
    direction = pos['direction']
    diff = close_price - pos['open_price'] if direction == "BUY" else pos['open_price'] - close_price
    pnl = round(diff * FIXED_LOT * 100, 2)  # XAU: 0.1lot = $1/pip, diff*100 = pips
    
    state['capital'] = round(state['capital'] + pnl, 2)
    state['total_trades'] += 1
    if pnl > 0:
        state['wins'] += 1
    else:
        state['losses'] += 1
    
    trade = {
        "strategy": pos['strategy'],
        "direction": pos['direction'],
        "open_price": pos['open_price'],
        "close_price": close_price,
        "pnl": pnl,
        "open_time": pos['open_time'],
        "close_time": time_now,
        "reason": reason
    }
    state['trades'].append(trade)
    state['positions'].remove(pos)
    
    log_trade({
        'time': time_now, 'strategy': pos['strategy'], 'action': '平仓',
        'direction': pos['direction'], 'price': close_price,
        'lots': FIXED_LOT, 'pnl': pnl, 'reason': reason
    })
    return pnl


# ── Excel 报表生成 ───────────────────────────────────────

def generate_excel():
    """读取CSV数据生成Excel报表"""
    wb = Workbook()
    
    # 样式
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    green_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    red_fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    def style_header(ws, row=1, max_col=30):
        for col in range(1, max_col+1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
    
    def style_data(ws, row, max_col=30):
        for col in range(1, max_col+1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = cell_align
            cell.border = thin_border
    
    # ─── Sheet 1: 交易明细 ───
    ws1 = wb.active
    ws1.title = "交易明细"
    ws1.append(['时间', '策略', '动作', '方向', '价格', '手数', '盈亏($)', '原因'])
    style_header(ws1, max_col=8)
    
    if os.path.exists(TRADE_LOG_FILE):
        with open(TRADE_LOG_FILE, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            headers = next(reader, None)
            for row_data in reader:
                ws1.append(row_data)
        
        # 样式和颜色
        for row in range(2, ws1.max_row + 1):
            style_data(ws1, row, max_col=8)
            pnl_cell = ws1.cell(row=row, column=7)
            try:
                pnl_val = float(pnl_cell.value or 0)
                if pnl_val > 0:
                    pnl_cell.fill = green_fill
                    pnl_cell.font = Font(color='006100')
                elif pnl_val < 0:
                    pnl_cell.fill = red_fill
                    pnl_cell.font = Font(color='9C0006')
            except:
                pass
    
    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 14
    ws1.column_dimensions['C'].width = 8
    ws1.column_dimensions['D'].width = 8
    ws1.column_dimensions['E'].width = 12
    ws1.column_dimensions['F'].width = 8
    ws1.column_dimensions['G'].width = 10
    ws1.column_dimensions['H'].width = 30
    
    # ─── Sheet 2: 因子评分 ───
    ws2 = wb.create_sheet("因子评分")
    
    if os.path.exists(SCORE_LOG_FILE):
        with open(SCORE_LOG_FILE, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            csv_headers = next(reader, None)
            if csv_headers:
                ws2.append(csv_headers)
                style_header(ws2, max_col=len(csv_headers))
                for row_data in reader:
                    ws2.append(row_data)
        
        for row in range(2, ws2.max_row + 1):
            style_data(ws2, row, max_col=len(csv_headers) if csv_headers else 29)
            # 信号列上色
            if csv_headers:
                for col_idx, h in enumerate(csv_headers):
                    if '信号' in h or '评分' in h:
                        cell = ws2.cell(row=row, column=col_idx+1)
                        val = str(cell.value or '')
                        if 'BUY' in val:
                            cell.fill = green_fill
                            cell.font = Font(color='006100', bold=True)
                        elif 'SELL' in val:
                            cell.fill = red_fill
                            cell.font = Font(color='9C0006', bold=True)
                        elif val == '3':
                            cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    else:
        ws2.append(['暂无因子评分数据，等待首次扫描...'])
    
    ws2.column_dimensions['A'].width = 20
    for col_letter in ['B', 'C', 'E', 'F', 'H', 'I', 'K', 'L', 'N', 'O', 'Q', 'R', 'S', 'U', 'V', 'W', 'X', 'Y', 'Z']:
        if col_letter in ws2.column_dimensions:
            ws2.column_dimensions[col_letter].width = 12
    # 说明列宽
    if ws2.max_column >= 5:
        ws2.column_dimensions['E'].width = 40
    if ws2.max_column >= 11:
        ws2.column_dimensions['K'].width = 40
    if ws2.max_column >= 17:
        ws2.column_dimensions['Q'].width = 40
    if ws2.max_column >= 23:
        ws2.column_dimensions['W'].width = 50
    
    # ─── Sheet 3: 策略表现 ───
    state = load_state()
    ws3 = wb.create_sheet("策略表现")
    
    # 统计
    strat_stats = {}
    for t in state.get('trades', []):
        s = t['strategy']
        if s not in strat_stats:
            strat_stats[s] = {'trades': 0, 'wins': 0, 'pnl': 0, 'max_dd': 0}
        strat_stats[s]['trades'] += 1
        strat_stats[s]['pnl'] += t['pnl']
        if t['pnl'] > 0:
            strat_stats[s]['wins'] += 1
    
    ws3.append(['策略', '交易次数', '胜率(%)', '总盈亏($)', '平均盈亏($)', '胜场', '负场'])
    style_header(ws3, max_col=7)
    
    total_pnl = 0
    for s_name in ['EMA_CROSS', 'RSI_REV', 'BB_SQUEEZE', 'ADX_RIDER']:
        stats = strat_stats.get(s_name, {'trades': 0, 'wins': 0, 'pnl': 0})
        win_rate = round(stats['wins'] / stats['trades'] * 100, 1) if stats['trades'] > 0 else 0
        avg_pnl = round(stats['pnl'] / stats['trades'], 2) if stats['trades'] > 0 else 0
        losses = stats['trades'] - stats['wins']
        row_data = [s_name, stats['trades'], win_rate, round(stats['pnl'], 2), avg_pnl, stats['wins'], losses]
        ws3.append(row_data)
        total_pnl += stats['pnl']
        row_idx = ws3.max_row
        style_data(ws3, row_idx, max_col=7)
    
    # 合计行
    total_trades = sum(s['trades'] for s in strat_stats.values())
    total_wins = sum(s['wins'] for s in strat_stats.values())
    ws3.append(['合计', total_trades, round(total_wins/total_trades*100, 1) if total_trades > 0 else 0,
                round(total_pnl, 2), '', total_wins, total_trades - total_wins])
    row_idx = ws3.max_row
    for col in range(1, 8):
        cell = ws3.cell(row=row_idx, column=col)
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.alignment = cell_align
    
    ws3.column_dimensions['A'].width = 16
    for col in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws3.column_dimensions[col].width = 14
    
    # ─── Sheet 4: 账户概览 ───
    ws4 = wb.create_sheet("账户概览")
    
    floating = 0
    for p in state.get('positions', []):
        diff = (state.get('_last_price', 0) - p['open_price']) if p['direction'] == "BUY" else (p['open_price'] - state.get('_last_price', 0))
        floating += diff * FIXED_LOT * 100
    
    ws4.append(['项目', '数值'])
    style_header(ws4, max_col=2)
    
    overview = [
        ('初始资金', f'${INITIAL_CAPITAL:.2f}'),
        ('当前资金', f'${state["capital"]:.2f}'),
        ('浮动盈亏', f'${floating:.2f}'),
        ('模拟权益', f'${state["capital"] + floating:.2f}'),
        ('总交易数', state['total_trades']),
        ('胜场', state['wins']),
        ('负场', state['losses']),
        ('胜率', f'{round(state["wins"]/state["total_trades"]*100, 1) if state["total_trades"] > 0 else 0}%'),
        ('当前持仓数', len(state['positions'])),
        ('每笔手数', f'{FIXED_LOT} 手'),
        ('更新时间', datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    
    for item, val in overview:
        ws4.append([item, val])
        style_data(ws4, ws4.max_row, max_col=2)
    
    ws4.column_dimensions['A'].width = 16
    ws4.column_dimensions['B'].width = 18
    
    # 最后更新原始数据
    ws4.append([])
    ws4.append(['当前持仓明细'])
    style_header(ws4, max_col=5)
    ws4.cell(row=ws4.max_row, column=1).font = Font(bold=True, color='FFFFFF')
    
    if state.get('positions'):
        ws4.append(['策略', '方向', '开仓价', '开仓时间', '浮动盈亏'])
        style_header(ws4, max_col=5)
        for p in state['positions']:
            diff = (state.get('_last_price', 0) - p['open_price']) if p['direction'] == "BUY" else (p['open_price'] - state.get('_last_price', 0))
            fl = round(diff * FIXED_LOT * 100, 2)
            ws4.append([p['strategy'], p['direction'], p['open_price'], p['open_time'], f'${fl:.2f}'])
            style_data(ws4, ws4.max_row, max_col=5)
    
    # 保存
    wb.save(EXCEL_REPORT)
    return EXCEL_REPORT


# ── 主循环 ───────────────────────────────────────────────

def main():
    import urllib.request
    
    # 1. Fetch data
    try:
        url_m30 = f"{ENGINE_URL}/api/market/candles?timeframe=M30&count=200"
        url_h1 = f"{ENGINE_URL}/api/market/candles?timeframe=H1&count=100"
        url_price = f"{ENGINE_URL}/api/market/price"
        
        req_m30 = urllib.request.Request(url_m30)
        with urllib.request.urlopen(req_m30, timeout=10) as resp:
            data_m30 = json.loads(resp.read().decode())
        
        req_h1 = urllib.request.Request(url_h1)
        with urllib.request.urlopen(req_h1, timeout=10) as resp:
            data_h1 = json.loads(resp.read().decode())
        
        req_price = urllib.request.Request(url_price)
        with urllib.request.urlopen(req_price, timeout=10) as resp:
            current_price = json.loads(resp.read().decode())
    except Exception as e:
        return f"[ERROR] 数据获取失败: {e}"
    
    # Dedup
    seen = {}
    for b in data_m30:
        seen[b['time']] = b
    m30 = sorted(seen.values(), key=lambda x: x['time'])
    seen_h1 = {}
    for b in data_h1:
        seen_h1[b['time']] = b
    h1 = sorted(seen_h1.values(), key=lambda x: x['time'])
    
    if len(m30) < 50 or len(h1) < 30:
        return f"[ERROR] 数据不足: M30={len(m30)}, H1={len(h1)}"
    
    bid = current_price.get('bid', m30[-1]['close'])
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # 2. Compute factor scores
    factor_data = compute_factors(m30, h1, current_price)
    if not factor_data:
        return "[ERROR] 因子计算失败"
    
    # 3. Log factor scores
    log_factor_scores(factor_data)
    f = factor_data['factors']
    
    # 4. Manage simulation positions
    state = load_state()
    state['_last_price'] = bid
    
    report = f"""
╔══════════════════════════════════════════════════╗
║  XAUUSD 手动策略模拟报告  {now_str}  ║
╚══════════════════════════════════════════════════╝

当前金价: {bid:.2f}
模拟资金: ${state['capital']:.2f} | 交易总数: {state['total_trades']}
胜率: {round(state['wins']/state['total_trades']*100, 1) if state['total_trades']>0 else 0}%
胜 {state['wins']} / 负 {state['losses']}

── 实时信号 ──"""

    for name, scores in f.items():
        report += f"\n  {name:<12} → 信号:{str(scores['信号']):>5}  评分:{scores['评分']}  {scores['说明']}"
    
    report += f"""

── 当前模拟持仓 ──"""
    if state['positions']:
        for p in state['positions']:
            dir_arrow = "▲" if p['direction'] == "BUY" else "▼"
            diff = (bid - p['open_price']) if p['direction'] == "BUY" else (p['open_price'] - bid)
            fl = round(diff * FIXED_LOT * 100, 2)
            report += f"\n  {dir_arrow} {p['strategy']} {p['direction']} @ {p['open_price']}  浮盈: ${fl:.2f}"
    else:
        report += "\n  无持仓"
    
    # 5. Check SL/TP and close conditions
    # Since we don't set SL/TP in this version, check for EXIT signals
    active_strats = {p['strategy'] for p in state['positions'] if p['strategy'] in f}
    
    for pos in list(state['positions']):
        strat_name = pos['strategy']
        if strat_name in f and f[strat_name]['信号'] == 'EXIT':
            pnl = close_position(state, pos, bid, f"信号出场: {f[strat_name]['说明']}", now_str)
            report += f"\n  📴 平仓 {strat_name} {pos['direction']} @ {bid:.2f} PnL=${pnl:.2f}"
    
    # 6. Check new entry signals
    for name, scores in f.items():
        if scores['信号'] in ('BUY', 'SELL') and name not in active_strats:
            # Check no opposite position exists
            has_opposite = any(p['strategy'] == name for p in state['positions'])
            if not has_opposite:
                open_position(state, name, scores['信号'], bid, now_str)
                arrow = "▲" if scores['信号'] == "BUY" else "▼"
                report += f"\n  🆕 {arrow} 开仓 {name} {scores['信号']} @ {bid:.2f}"
    
    # 7. Update equity
    floating = 0
    for p in state['positions']:
        diff = (bid - p['open_price']) if p['direction'] == "BUY" else (p['open_price'] - bid)
        floating += diff * FIXED_LOT * 100
    state['equity'] = round(state['capital'] + floating, 2)
    save_state(state)
    
    # 8. Check if anything changed (quiet mode support)
    had_actions = bool([l for l in report.split('\n') if '🆕' in l or '📴' in l or '🔴' in l or '🟢' in l])
    
    # 9. Generate Excel
    excel_path = generate_excel()
    
    # Quiet mode: suppress output when nothing happened (cron watchdog pattern)
    quiet = os.environ.get('SIM_QUIET', '0') == '1'
    if quiet and not had_actions:
        return ""  # Silent - nothing new
    
    report += f"""
  模拟权益: ${state['equity']:.2f} (浮动: ${floating:.2f})
  Excel报表: {excel_path}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"""
    
    return report


if __name__ == '__main__':
    result = main()
    print(result)
