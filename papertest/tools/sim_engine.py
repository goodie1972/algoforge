#!/c/Python314/python
"""
XAUUSD 手动策略模拟系统 v2.0 - Excel 完整报表版
跑在实盘信号上，不执行操作。每15分钟检查一次。
"""

import json, os, sys, math, time, csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers

# ── 配置 ────────────────────────────────────────────────
INITIAL_CAPITAL = 10000.0      # 模拟初始资金 USD
FIXED_LOT = 0.1                # 固定每笔 0.1手
ENGINE_URL = "http://localhost:1783"
DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
os.makedirs(DATA_DIR, exist_ok=True)

STATE_FILE = os.path.join(DATA_DIR, "sim_state.json")
SCORE_LOG_FILE = os.path.join(DATA_DIR, "factor_scores.csv")
TRADE_LOG_FILE = os.path.join(DATA_DIR, "trade_log.csv")
EXCEL_REPORT = os.path.join(os.path.dirname(__file__), "模拟交易报表.xlsx")


# ── 指标计算 ────────────────────────────────────────────

def ema(data, period):
    """计算EMA"""
    if len(data) < period:
        return data
    k = 2.0 / (period + 1)
    ema_val = sum(data[:period]) / period
    result = [ema_val] * (period - 1)  # pad
    result.append(ema_val)
    for price in data[period:]:
        ema_val = price * k + ema_val * (1 - k)
        result.append(round(ema_val, 2))
    return result

def rsi(closes, period=14):
    """计算RSI"""
    if len(closes) < period + 1:
        return [50] * len(closes)
    gains, losses = [], []
    for i in range(1, len(closes)):
        diff = closes[i] - closes[i-1]
        gains.append(max(diff, 0))
        losses.append(max(-diff, 0))
    result = [50] * period
    avg_gain = sum(gains[:period]) / period
    avg_loss = sum(losses[:period]) / period
    for i in range(period, len(gains)):
        if avg_loss == 0:
            result.append(100.0)
        else:
            rs = avg_gain / avg_loss
            result.append(round(100 - 100 / (1 + rs), 1))
        avg_gain = (avg_gain * (period - 1) + gains[i]) / period
        avg_loss = (avg_loss * (period - 1) + losses[i]) / period
    return result

def sma(data, period):
    if len(data) < period:
        return [sum(data)/len(data)] * len(data)
    result = [0] * len(data)
    cum = sum(data[:period])
    result[period-1] = round(cum / period, 2)
    for i in range(period, len(data)):
        cum = cum - data[i-period] + data[i]
        result[i] = round(cum / period, 2)
    for i in range(period-1):
        result[i] = result[period-1]
    return result

def bollinger(closes, period=20, std_dev=2):
    """Returns (middle, upper, lower, bandwidth) arrays"""
    ma = sma(closes, period)
    upper, lower, bandwidth = [], [], []
    for i in range(len(closes)):
        if i < period-1:
            upper.append(round(ma[i], 2))
            lower.append(round(ma[i], 2))
            bandwidth.append(0)
            continue
        window = closes[i-period+1:i+1]
        sd = (sum((x - ma[i])**2 for x in window) / period) ** 0.5
        upper.append(round(ma[i] + std_dev * sd, 2))
        lower.append(round(ma[i] - std_dev * sd, 2))
        bw = (upper[-1] - lower[-1]) / ma[i] * 100 if ma[i] > 0 else 0
        bandwidth.append(round(bw, 2))
    return ma, upper, lower, bandwidth

def atr(candles, period=14):
    """计算ATR"""
    if len(candles) < period + 1:
        return 0
    trs = []
    for i in range(1, len(candles)):
        high, low = candles[i]['high'], candles[i]['low']
        prev_close = candles[i-1]['close']
        tr = max(high-low, abs(high-prev_close), abs(low-prev_close))
        trs.append(tr)
    return sum(trs[-period:]) / period if trs else 0


# ── 因子评分系统 v3.0 — 高级多因子融合 ──────────────────

def adx(candles, period=14):
    """计算ADX, +DI, -DI"""
    if len(candles) < period + 1:
        return [0]*len(candles), [0]*len(candles), [0]*len(candles)
    trs, plus_dms, minus_dms = [], [], []
    for i in range(1, len(candles)):
        high, low = candles[i]['high'], candles[i]['low']
        prev_close = candles[i-1]['close']
        tr = max(high-low, abs(high-prev_close), abs(low-prev_close))
        trs.append(tr)
        up_move = high - candles[i-1]['high']
        down_move = candles[i-1]['low'] - low
        plus_dms.append(max(up_move, 0) if up_move > down_move else 0)
        minus_dms.append(max(down_move, 0) if down_move > up_move else 0)
    def smooth(values, n):
        result = [0]*len(values)
        if len(values) < n:
            return result
        k = 1.0/n
        smoothed = sum(values[:n])/n
        result[n-1] = smoothed
        for i in range(n, len(values)):
            smoothed = (smoothed*(n-1)+values[i])/n
            result[i] = smoothed
        return result
    tr_ema = smooth(trs, period)
    pdi_ema = smooth(plus_dms, period)
    ndi_ema = smooth(minus_dms, period)
    pad_len = len(candles) - len(pdi_ema)
    dx_vals = []
    for i in range(len(pdi_ema)):
        di_sum = pdi_ema[i]+ndi_ema[i]
        dx_vals.append(abs(pdi_ema[i]-ndi_ema[i])/di_sum*100 if di_sum > 0 else 0)
    adx_raw = smooth(dx_vals, period)
    max_val = max(max(tr_ema), 1)
    pdi_pct = [v/max_val*100 if max_val > 0 else 0 for v in pdi_ema]
    ndi_pct = [v/max_val*100 if max_val > 0 else 0 for v in ndi_ema]
    return [0]*pad_len+adx_raw, [0]*pad_len+pdi_pct, [0]*pad_len+ndi_pct

def macd(closes, fast=12, slow=26, signal=9):
    """计算MACD线、信号线、柱状图"""
    if len(closes) < slow + signal:
        return [0]*len(closes), [0]*len(closes), [0]*len(closes)
    ema_f = ema(closes, fast)
    ema_s = ema(closes, slow)
    macd_line = [ema_f[i]-ema_s[i] for i in range(len(closes))]
    sig_line = [0]*len(closes)
    k = 2.0/(signal+1)
    sig_line[slow+signal-1] = sum(macd_line[slow:slow+signal])/signal
    for i in range(slow+signal, len(closes)):
        sig_line[i] = macd_line[i]*k + sig_line[i-1]*(1-k)
    hist = [macd_line[i]-sig_line[i] for i in range(len(closes))]
    return macd_line, sig_line, hist

def heikin_ashi(candles):
    """将普通K线转换为Heikin Ashi"""
    ha = []
    for i, c in enumerate(candles):
        if i == 0:
            ha_o = c['open']
        else:
            ha_o = (ha[-1]['open']+ha[-1]['close'])/2
        ha_c = (c['open']+c['high']+c['low']+c['close'])/4
        ha_h = max(c['high'], ha_o, ha_c)
        ha_l = min(c['low'], ha_o, ha_c)
        ha.append({'time':c['time'],'open':ha_o,'high':ha_h,'low':ha_l,'close':ha_c})
    return ha

def detect_swing_highs_lows(candles, lookback=5):
    """检测摆动高点和低点"""
    highs = [c['high'] for c in candles]
    lows = [c['low'] for c in candles]
    swing_highs, swing_lows = set(), set()
    for i in range(lookback, len(candles)-lookback):
        if all(highs[i] >= highs[i-j] for j in range(1, lookback+1)) and \
           all(highs[i] >= highs[i+j] for j in range(1, lookback+1)):
            swing_highs.add(i)
        if all(lows[i] <= lows[i-j] for j in range(1, lookback+1)) and \
           all(lows[i] <= lows[i+j] for j in range(1, lookback+1)):
            swing_lows.add(i)
    return swing_highs, swing_lows

def compute_factors(m30, h1, curr_price):
    """计算v3高级策略因子"""
    if len(m30) < 50 or len(h1) < 30:
        return None
    
    closes_m30 = [b['close'] for b in m30]
    closes_h1 = [b['close'] for b in h1]
    bid = curr_price.get('bid', m30[-1]['close'])
    now_ts = m30[-1]['time']
    now_str = datetime.fromtimestamp(now_ts).strftime("%Y-%m-%d %H:%M:%S")
    
    # ─── 公共指标 ───
    ema9_h1 = ema(closes_h1, 9)
    ema21_h1 = ema(closes_h1, 21)
    ema50_h1 = ema(closes_h1, 50)
    ema9_curr, ema21_curr, ema50_curr = ema9_h1[-1], ema21_h1[-1], ema50_h1[-1]
    
    rsi_vals = rsi(closes_m30, 14)
    rsi_curr = rsi_vals[-1]
    
    bb_mid, bb_upper, bb_lower, bb_bw = bollinger(closes_m30, 20, 2)
    bb_mid_c, bb_up_c, bb_lo_c, bb_bw_c = bb_mid[-1], bb_upper[-1], bb_lower[-1], bb_bw[-1]
    bb_bw_avg = sum(bb_bw[-20:]) / 20 if len(bb_bw) >= 20 else bb_bw_c
    bb_squeeze = bb_bw_c < bb_bw_avg * 0.7 if bb_bw_avg > 0 else False
    
    adx_vals, pdi_vals, ndi_vals = adx(h1, 14)
    adx_c, pdi_c, ndi_c = adx_vals[-1], pdi_vals[-1], ndi_vals[-1]
    
    macd_l, macd_s, macd_h = macd(closes_h1, 12, 26, 9)
    macd_curr, macd_sig, macd_hist = macd_l[-1], macd_s[-1], macd_h[-1]
    macd_hist_prev = macd_h[-2] if len(macd_h) >= 2 else 0
    
    atr_val = atr(m30, 14)
    
    # Heikin Ashi
    ha_m30 = heikin_ashi(m30[-10:])
    ha_bull = ha_m30[-1]['close'] > ha_m30[-1]['open'] if ha_m30 else False
    ha_bear = ha_m30[-1]['close'] < ha_m30[-1]['open'] if ha_m30 else False
    
    # 摆动点
    sh, sl = detect_swing_highs_lows(h1, 5)
    h1_highs = [h1[i]['high'] for i in sh if i < len(h1)]
    h1_lows = [h1[i]['low'] for i in sl if i < len(h1)]
    
    # 趋势结构
    trend_up_h1 = ema9_curr > ema21_curr > ema50_curr
    trend_down_h1 = ema9_curr < ema21_curr < ema50_curr
    ema_slope_h1 = (ema9_h1[-1] - ema9_h1[-5]) / 5 if len(ema9_h1) >= 5 else 0
    
    # 20-bar 高低点
    recent_high = max(h['high'] for h in h1[-20:])
    recent_low = min(h['low'] for h in h1[-20:])
    breakout_up = bid > recent_high
    breakout_dn = bid < recent_low
    
    # 日内时段 (UTC+8)
    hour = datetime.now().hour
    asia_session = 6 <= hour < 14
    london_session = 15 <= hour < 23
    ny_session = 21 <= hour or hour < 5
    best_session = london_session or ny_session
    
    # 初始平衡范围 (今日前3根H1)
    h1_today = [b for b in h1 if b['time'] > now_ts - 86400]
    ib_high = max(b['high'] for b in h1_today[:3]) if len(h1_today) >= 3 else 0
    ib_low = min(b['low'] for b in h1_today[:3]) if len(h1_today) >= 3 else 0
    ib_range = ib_high - ib_low if ib_high > ib_low else 0
    
    factors = {}
    
    # ═══════════════════════════════════════════════════
    #  策略1: MULTI_CONFLUENCE (8因子评分)
    #  来源: FMZ 10-Point Scoring System 改编
    #  核心理念: 多因子汇聚评分, 6/8分入场
    # ═══════════════════════════════════════════════════
    mc = {}
    score = 0
    reasons = []
    
    if trend_up_h1 or trend_down_h1:
        score += 1; reasons.append('EMA趋势')
    if rsi_curr < 35 or rsi_curr > 65:
        score += 1; reasons.append('RSI极值')
    if (macd_hist > 0 and macd_hist_prev <= 0) or (macd_hist < 0 and macd_hist_prev >= 0):
        score += 1; reasons.append('MACD交叉')
    if bid <= bb_lo_c*1.01 or bid >= bb_up_c*0.99:
        score += 1; reasons.append('BB外轨')
    last_body = abs(m30[-1]['close']-m30[-1]['open'])
    avg_body = sum(abs(m30[i]['close']-m30[i]['open']) for i in range(-10,0))/10
    if last_body > avg_body*1.5:
        score += 1; reasons.append('量能扩张')
    if len(h1_highs)>=2 and h1_highs[-1]>h1_highs[-2] or len(h1_lows)>=2 and h1_lows[-1]<h1_lows[-2]:
        score += 1; reasons.append('结构突破')
    body_ratio = abs(m30[-1]['close']-m30[-1]['open'])/(m30[-1]['high']-m30[-1]['low']) if m30[-1]['high']-m30[-1]['low']>0 else 0
    if body_ratio > 0.7:
        score += 1; reasons.append('强势K线')
    if best_session:
        score += 1; reasons.append('最佳时段')
    
    mc['评分'] = score
    mc['总分'] = f'{score}/8'
    mc['因子明细'] = ' | '.join(reasons)
    
    direction = ''
    if score >= 6:
        bull = sum([trend_up_h1, rsi_curr<35, macd_hist>0, bid<=bb_lo_c*1.01, ha_bull, breakout_up])
        bear = sum([trend_down_h1, rsi_curr>65, macd_hist<0, bid>=bb_up_c*0.99, ha_bear, breakout_dn])
        direction = 'BUY' if bull >= bear else 'SELL'
        mc['信号'] = direction
        mc['说明'] = f'强信号({score}/8)→{direction}'
    elif score >= 4 and breakout_up and trend_up_h1:
        mc['信号'] = 'BUY'; direction = 'BUY'
        mc['说明'] = f'弱信号({score}/8)+突破→BUY'
    elif score >= 4 and breakout_dn and trend_down_h1:
        mc['信号'] = 'SELL'; direction = 'SELL'
        mc['说明'] = f'弱信号({score}/8)+跌破→SELL'
    else:
        mc['信号'] = '无'
        mc['说明'] = f'评分{score}/8低于阈值'
    mc['EMA9'], mc['EMA21'], mc['RSI'] = ema9_curr, ema21_curr, rsi_curr
    mc['MACD柱'] = round(macd_hist,2)
    mc['量能比'] = round(last_body/avg_body, 2) if avg_body > 0 else 1
    factors['MULTI_CONFLUENCE'] = mc
    
    # ═══════════════════════════════════════════════════
    #  策略2: MOMENTUM_BREAK (趋势+突破)
    #  来源: Gold Momentum Edge v9 + 结构突破
    #  核心理念: 趋势方向过滤, ADX强度确认, 结构突破入场
    # ═══════════════════════════════════════════════════
    mb = {}
    mb['EMA50'] = ema50_curr
    mb['ADX'] = round(adx_c, 1)
    mb['20高'], mb['20低'] = round(recent_high,2), round(recent_low,2)
    mb['H1趋势'] = '↑多头' if trend_up_h1 else ('↓空头' if trend_down_h1 else '→横盘')
    
    score_mb = 0
    if trend_up_h1 and adx_c > 25 and breakout_up:
        score_mb = 3
        mb['信号'] = 'BUY'
        mb['说明'] = f'多头+ADX({adx_c:.1f})强趋势+突破{recent_high:.1f}高点'
    elif trend_down_h1 and adx_c > 25 and breakout_dn:
        score_mb = 3
        mb['信号'] = 'SELL'
        mb['说明'] = f'空头+ADX({adx_c:.1f})强趋势+跌破{recent_low:.1f}低点'
    elif trend_up_h1 and not breakout_up:
        mb['信号'] = '无'
        mb['说明'] = f'多头趋势但未突破(距高点{round(recent_high-bid,2)})'
    elif trend_down_h1 and not breakout_dn:
        mb['信号'] = '无'
        mb['说明'] = f'空头趋势但未跌破(距低点{round(bid-recent_low,2)})'
    elif adx_c < 20:
        mb['信号'] = 'EXIT'
        mb['说明'] = f'ADX({adx_c:.1f})<20趋势消失'
    else:
        mb['信号'] = '无'
        mb['说明'] = '条件未满足'
    mb['评分'] = score_mb
    factors['MOMENTUM_BREAK'] = mb
    
    # ═══════════════════════════════════════════════════
    #  策略3: HA_STOCHASTIC (Heikin Ashi + 随机)
    #  来源: Heikin Ashi Scalper XAU + Gold G v1.1
    #  核心理念: HA滤噪音, RSI极端 + BB外轨 = 反转入场
    # ═══════════════════════════════════════════════════
    hs = {}
    hs['HA方向'] = '↑' if ha_bull else ('↓' if ha_bear else '→')
    hs['RSI'] = rsi_curr
    hs['BB上'], hs['BB下'] = round(bb_up_c,2), round(bb_lo_c,2)
    hs['ATR'] = round(atr_val, 2)
    hs['挤仓'] = '是' if bb_squeeze else '否'
    
    stoch_k = (bid-bb_lo_c)/(bb_up_c-bb_lo_c)*100 if bb_up_c>bb_lo_c else 50
    hs['随机K'] = round(stoch_k, 1)
    
    score_hs = 0
    if ha_bull and rsi_curr < 30 and bid <= bb_lo_c*1.01:
        score_hs = 3
        hs['信号'] = 'BUY'
        hs['说明'] = f'HA多头+RSI超卖({rsi_curr})+BB下轨'
    elif ha_bear and rsi_curr > 70 and bid >= bb_up_c*0.99:
        score_hs = 3
        hs['信号'] = 'SELL'
        hs['说明'] = f'HA空头+RSI超买({rsi_curr})+BB上轨'
    elif bb_squeeze and ha_bull and breakout_up:
        score_hs = 2
        hs['信号'] = 'BUY'
        hs['说明'] = 'BB挤仓+HA多头+突破'
    elif bb_squeeze and ha_bear and breakout_dn:
        score_hs = 2
        hs['信号'] = 'SELL'
        hs['说明'] = 'BB挤仓+HA空头+跌破'
    elif ha_bull and rsi_curr > 60:
        hs['信号'] = '无'; hs['说明'] = f'HA多头但RSI偏高({rsi_curr})不追'
    elif ha_bear and rsi_curr < 40:
        hs['信号'] = '无'; hs['说明'] = f'HA空头但RSI偏低({rsi_curr})不杀'
    else:
        hs['信号'] = '无'
        hs['说明'] = f'HA{"多头" if ha_bull else "空头"}RSI={rsi_curr}'
    hs['评分'] = score_hs
    factors['HA_STOCHASTIC'] = hs
    
    # ═══════════════════════════════════════════════════
    #  策略4: IB_BREAKOUT (初始平衡突破)
    #  来源: NY Open Range Breakout / Initial Balance
    #  核心理念: 前3根H1形成IB, 突破后回踩入场
    # ═══════════════════════════════════════════════════
    ib = {}
    ib['IB高'] = round(ib_high, 2) if ib_high else 0
    ib['IB低'] = round(ib_low, 2) if ib_low else 0
    ib['IB幅度'] = round(ib_range, 2) if ib_range else 0
    ib['时段'] = '亚盘' if asia_session else ('伦敦' if london_session else '纽约')
    
    score_ib = 0
    if ib_range > 0 and ib_range < 50:  # 硬上限50点, 防止极值日
        if bid > ib_high:
            retrace_entry = ib_high + ib_range*0.25
            if bid <= retrace_entry:
                score_ib = 3
                ib['信号'] = 'BUY'
                ib['说明'] = f'突破IB高({round(ib_high,2)})回踩{round(retrace_entry,2)}入场'
            else:
                ib['信号'] = '无'
                ib['说明'] = f'突破IB高但已远离, etc'
        elif bid < ib_low:
            retrace_entry = ib_low - ib_range*0.25
            if bid >= retrace_entry:
                score_ib = 3
                ib['信号'] = 'SELL'
                ib['说明'] = f'跌破IB低({round(ib_low,2)})反弹{round(retrace_entry,2)}入场'
            else:
                ib['信号'] = '无'
                ib['说明'] = f'跌破IB低但已远离'
        else:
            ib['信号'] = '无'
            ib['说明'] = f'IB内({round(ib_low,2)}-{round(ib_high,2)})等待突破'
    elif ib_range == 0:
        ib['信号'] = '无'
        ib['说明'] = '数据不足计算IB'
    else:
        ib['信号'] = '无'
        ib['说明'] = f'IB幅度({round(ib_range,2)})过大'
    ib['评分'] = score_ib
    if score_ib == 3 and ib['信号'] == 'BUY':
        ib['SL'] = round(ib_high-ib_range*0.6, 2)
        ib['TP'] = round(ib_high+ib_range*0.5, 2)
    elif score_ib == 3 and ib['信号'] == 'SELL':
        ib['SL'] = round(ib_low+ib_range*0.6, 2)
        ib['TP'] = round(ib_low-ib_range*0.5, 2)
    else:
        ib['SL'] = ib['TP'] = 0
    factors['IB_BREAKOUT'] = ib
    
    return {
        'time': now_str,
        'timestamp': now_ts,
        'price': round(bid, 2),
        'factors': factors
    }


# ── 模拟仓管理 ───────────────────────────────────────────

def load_state():
    if os.path.exists(STATE_FILE):
        with open(STATE_FILE) as f:
            return json.load(f)
    return {
        "capital": INITIAL_CAPITAL,
        "equity": INITIAL_CAPITAL,
        "positions": [],
        "trades": [],
        "total_trades": 0,
        "wins": 0,
        "losses": 0
    }

def save_state(state):
    with open(STATE_FILE, 'w') as f:
        json.dump(state, f, indent=2)

def log_factor_scores(factor_data):
    """追加v3高级策略因子评分到CSV"""
    file_exists = os.path.exists(SCORE_LOG_FILE)
    with open(SCORE_LOG_FILE, 'a', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        if not file_exists:
            headers = [
                '时间', '价格',
                'MC_信号', 'MC_评分', 'MC_总分', 'MC_说明', 'MC_EMA9', 'MC_EMA21', 'MC_RSI', 'MC_MACD柱', 'MC_量能比', 'MC_因子明细',
                'MB_信号', 'MB_评分', 'MB_说明', 'MB_EMA50', 'MB_ADX', 'MB_20高', 'MB_20低', 'MB_H1趋势',
                'HS_信号', 'HS_评分', 'HS_说明', 'HS_HA方向', 'HS_RSI', 'HS_BB上', 'HS_BB下', 'HS_随机K', 'HS_ATR', 'HS_挤仓',
                'IB_信号', 'IB_评分', 'IB_说明', 'IB_IB高', 'IB_IB低', 'IB_IB幅度', 'IB_时段', 'IB_SL', 'IB_TP'
            ]
            writer.writerow(headers)
        
        f = factor_data['factors']
        mc, mb, hs, ib = f['MULTI_CONFLUENCE'], f['MOMENTUM_BREAK'], f['HA_STOCHASTIC'], f['IB_BREAKOUT']
        row = [
            factor_data['time'], factor_data['price'],
            mc['信号'], mc['评分'], mc['总分'], mc['说明'], mc['EMA9'], mc['EMA21'], mc['RSI'], mc['MACD柱'], mc['量能比'], mc['因子明细'],
            mb['信号'], mb['评分'], mb['说明'], mb['EMA50'], mb['ADX'], mb['20高'], mb['20低'], mb['H1趋势'],
            hs['信号'], hs['评分'], hs['说明'], hs['HA方向'], hs['RSI'], hs['BB上'], hs['BB下'], hs['随机K'], hs['ATR'], hs['挤仓'],
            ib['信号'], ib['评分'], ib['说明'], ib['IB高'], ib['IB低'], ib['IB幅度'], ib['时段'], ib.get('SL',0), ib.get('TP',0)
        ]
        writer.writerow(row)

def log_trade(trade_data):
    """追加交易记录到CSV"""
    file_exists = os.path.exists(TRADE_LOG_FILE)
    with open(TRADE_LOG_FILE, 'a', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        if not file_exists:
            headers = ['时间', '策略', '动作', '方向', '价格', '手数', '盈亏($)', '原因']
            writer.writerow(headers)
        writer.writerow([
            trade_data['time'],
            trade_data['strategy'],
            trade_data['action'],
            trade_data.get('direction', ''),
            trade_data['price'],
            trade_data.get('lots', FIXED_LOT),
            trade_data.get('pnl', 0),
            trade_data.get('reason', '')
        ])


def open_position(state, strategy, direction, price, time_now):
    pos = {
        "strategy": strategy,
        "direction": direction,
        "open_price": price,
        "lots": FIXED_LOT,
        "open_time": time_now
    }
    state['positions'].append(pos)
    log_trade({
        'time': time_now, 'strategy': strategy, 'action': '开仓',
        'direction': direction, 'price': price, 'lots': FIXED_LOT,
        'pnl': 0, 'reason': f'{strategy}信号触发'
    })
    return pos

def close_position(state, pos, close_price, reason, time_now):
    direction = pos['direction']
    diff = close_price - pos['open_price'] if direction == "BUY" else pos['open_price'] - close_price
    pnl = round(diff * FIXED_LOT * 100, 2)  # XAU: 0.1lot = $1/pip, diff*100 = pips
    
    state['capital'] = round(state['capital'] + pnl, 2)
    state['total_trades'] += 1
    if pnl > 0:
        state['wins'] += 1
    else:
        state['losses'] += 1
    
    trade = {
        "strategy": pos['strategy'],
        "direction": pos['direction'],
        "open_price": pos['open_price'],
        "close_price": close_price,
        "pnl": pnl,
        "open_time": pos['open_time'],
        "close_time": time_now,
        "reason": reason
    }
    state['trades'].append(trade)
    state['positions'].remove(pos)
    
    log_trade({
        'time': time_now, 'strategy': pos['strategy'], 'action': '平仓',
        'direction': pos['direction'], 'price': close_price,
        'lots': FIXED_LOT, 'pnl': pnl, 'reason': reason
    })
    return pnl


# ── Excel 报表生成 ───────────────────────────────────────

def generate_excel():
    """读取CSV数据生成Excel报表"""
    wb = Workbook()
    
    # 样式
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    green_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    red_fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    def style_header(ws, row=1, max_col=30):
        for col in range(1, max_col+1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
    
    def style_data(ws, row, max_col=30):
        for col in range(1, max_col+1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = cell_align
            cell.border = thin_border
    
    # ─── Sheet 1: 交易明细 ───
    ws1 = wb.active
    ws1.title = "交易明细"
    ws1.append(['时间', '策略', '动作', '方向', '价格', '手数', '盈亏($)', '原因'])
    style_header(ws1, max_col=8)
    
    if os.path.exists(TRADE_LOG_FILE):
        with open(TRADE_LOG_FILE, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            headers = next(reader, None)
            for row_data in reader:
                ws1.append(row_data)
        
        # 样式和颜色
        for row in range(2, ws1.max_row + 1):
            style_data(ws1, row, max_col=8)
            pnl_cell = ws1.cell(row=row, column=7)
            try:
                pnl_val = float(pnl_cell.value or 0)
                if pnl_val > 0:
                    pnl_cell.fill = green_fill
                    pnl_cell.font = Font(color='006100')
                elif pnl_val < 0:
                    pnl_cell.fill = red_fill
                    pnl_cell.font = Font(color='9C0006')
            except:
                pass
    
    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 14
    ws1.column_dimensions['C'].width = 8
    ws1.column_dimensions['D'].width = 8
    ws1.column_dimensions['E'].width = 12
    ws1.column_dimensions['F'].width = 8
    ws1.column_dimensions['G'].width = 10
    ws1.column_dimensions['H'].width = 30
    
    # ─── Sheet 2: 因子评分 ───
    ws2 = wb.create_sheet("因子评分")
    
    if os.path.exists(SCORE_LOG_FILE):
        with open(SCORE_LOG_FILE, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            csv_headers = next(reader, None)
            if csv_headers:
                ws2.append(csv_headers)
                style_header(ws2, max_col=len(csv_headers))
                for row_data in reader:
                    ws2.append(row_data)
        
        for row in range(2, ws2.max_row + 1):
            style_data(ws2, row, max_col=len(csv_headers) if csv_headers else 29)
            # 信号列上色
            if csv_headers:
                for col_idx, h in enumerate(csv_headers):
                    if '信号' in h or '评分' in h:
                        cell = ws2.cell(row=row, column=col_idx+1)
                        val = str(cell.value or '')
                        if 'BUY' in val:
                            cell.fill = green_fill
                            cell.font = Font(color='006100', bold=True)
                        elif 'SELL' in val:
                            cell.fill = red_fill
                            cell.font = Font(color='9C0006', bold=True)
                        elif val == '3':
                            cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    else:
        ws2.append(['暂无因子评分数据，等待首次扫描...'])
    
    ws2.column_dimensions['A'].width = 20
    for col_letter in ['B', 'C', 'E', 'F', 'H', 'I', 'K', 'L', 'N', 'O', 'Q', 'R', 'S', 'U', 'V', 'W', 'X', 'Y', 'Z']:
        if col_letter in ws2.column_dimensions:
            ws2.column_dimensions[col_letter].width = 12
    # 说明列宽
    if ws2.max_column >= 5:
        ws2.column_dimensions['E'].width = 40
    if ws2.max_column >= 11:
        ws2.column_dimensions['K'].width = 40
    if ws2.max_column >= 17:
        ws2.column_dimensions['Q'].width = 40
    if ws2.max_column >= 23:
        ws2.column_dimensions['W'].width = 50
    
    # ─── Sheet 3: 策略表现 ───
    state = load_state()
    ws3 = wb.create_sheet("策略表现")
    
    # 统计
    strat_stats = {}
    for t in state.get('trades', []):
        s = t['strategy']
        if s not in strat_stats:
            strat_stats[s] = {'trades': 0, 'wins': 0, 'pnl': 0, 'max_dd': 0}
        strat_stats[s]['trades'] += 1
        strat_stats[s]['pnl'] += t['pnl']
        if t['pnl'] > 0:
            strat_stats[s]['wins'] += 1
    
    ws3.append(['策略', '交易次数', '胜率(%)', '总盈亏($)', '平均盈亏($)', '胜场', '负场'])
    style_header(ws3, max_col=7)
    
    total_pnl = 0
    for s_name in ['MULTI_CONFLUENCE', 'MOMENTUM_BREAK', 'HA_STOCHASTIC', 'IB_BREAKOUT']:
        stats = strat_stats.get(s_name, {'trades': 0, 'wins': 0, 'pnl': 0})
        win_rate = round(stats['wins'] / stats['trades'] * 100, 1) if stats['trades'] > 0 else 0
        avg_pnl = round(stats['pnl'] / stats['trades'], 2) if stats['trades'] > 0 else 0
        losses = stats['trades'] - stats['wins']
        row_data = [s_name, stats['trades'], win_rate, round(stats['pnl'], 2), avg_pnl, stats['wins'], losses]
        ws3.append(row_data)
        total_pnl += stats['pnl']
        row_idx = ws3.max_row
        style_data(ws3, row_idx, max_col=7)
    
    # 合计行
    total_trades = sum(s['trades'] for s in strat_stats.values())
    total_wins = sum(s['wins'] for s in strat_stats.values())
    ws3.append(['合计', total_trades, round(total_wins/total_trades*100, 1) if total_trades > 0 else 0,
                round(total_pnl, 2), '', total_wins, total_trades - total_wins])
    row_idx = ws3.max_row
    for col in range(1, 8):
        cell = ws3.cell(row=row_idx, column=col)
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.alignment = cell_align
    
    ws3.column_dimensions['A'].width = 16
    for col in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws3.column_dimensions[col].width = 14
    
    # ─── Sheet 4: 账户概览 ───
    ws4 = wb.create_sheet("账户概览")
    
    floating = 0
    for p in state.get('positions', []):
        diff = (state.get('_last_price', 0) - p['open_price']) if p['direction'] == "BUY" else (p['open_price'] - state.get('_last_price', 0))
        floating += diff * FIXED_LOT * 100
    
    ws4.append(['项目', '数值'])
    style_header(ws4, max_col=2)
    
    overview = [
        ('初始资金', f'${INITIAL_CAPITAL:.2f}'),
        ('当前资金', f'${state["capital"]:.2f}'),
        ('浮动盈亏', f'${floating:.2f}'),
        ('模拟权益', f'${state["capital"] + floating:.2f}'),
        ('总交易数', state['total_trades']),
        ('胜场', state['wins']),
        ('负场', state['losses']),
        ('胜率', f'{round(state["wins"]/state["total_trades"]*100, 1) if state["total_trades"] > 0 else 0}%'),
        ('当前持仓数', len(state['positions'])),
        ('每笔手数', f'{FIXED_LOT} 手'),
        ('更新时间', datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    
    for item, val in overview:
        ws4.append([item, val])
        style_data(ws4, ws4.max_row, max_col=2)
    
    ws4.column_dimensions['A'].width = 16
    ws4.column_dimensions['B'].width = 18
    
    # 最后更新原始数据
    ws4.append([])
    ws4.append(['当前持仓明细'])
    style_header(ws4, max_col=5)
    ws4.cell(row=ws4.max_row, column=1).font = Font(bold=True, color='FFFFFF')
    
    if state.get('positions'):
        ws4.append(['策略', '方向', '开仓价', '开仓时间', '浮动盈亏'])
        style_header(ws4, max_col=5)
        for p in state['positions']:
            diff = (state.get('_last_price', 0) - p['open_price']) if p['direction'] == "BUY" else (p['open_price'] - state.get('_last_price', 0))
            fl = round(diff * FIXED_LOT * 100, 2)
            ws4.append([p['strategy'], p['direction'], p['open_price'], p['open_time'], f'${fl:.2f}'])
            style_data(ws4, ws4.max_row, max_col=5)
    
    # 保存
    wb.save(EXCEL_REPORT)
    return EXCEL_REPORT


# ── 主循环 ───────────────────────────────────────────────

def main():
    import urllib.request
    
    # 1. Fetch data
    try:
        # 获取最近7天的K线（约336条M30 + 168条H1，远超所需）
        now = int(time.time())
        start_ts = now - 7 * 86400
        url_m30 = f"{ENGINE_URL}/api/data/candles?timeframe=M30&limit=250&start_ts={start_ts}"
        url_h1 = f"{ENGINE_URL}/api/data/candles?timeframe=H1&limit=150&start_ts={start_ts}"
        url_price = f"{ENGINE_URL}/api/market/price"
        
        req_m30 = urllib.request.Request(url_m30)
        with urllib.request.urlopen(req_m30, timeout=10) as resp:
            data_m30 = json.loads(resp.read().decode())
        
        req_h1 = urllib.request.Request(url_h1)
        with urllib.request.urlopen(req_h1, timeout=10) as resp:
            data_h1 = json.loads(resp.read().decode())
        
        req_price = urllib.request.Request(url_price)
        with urllib.request.urlopen(req_price, timeout=10) as resp:
            current_price = json.loads(resp.read().decode())
    except Exception as e:
        return f"[ERROR] 数据获取失败: {e}"
    
    # Dedup
    seen = {}
    for b in data_m30:
        seen[b['time']] = b
    m30 = sorted(seen.values(), key=lambda x: x['time'])
    seen_h1 = {}
    for b in data_h1:
        seen_h1[b['time']] = b
    h1 = sorted(seen_h1.values(), key=lambda x: x['time'])
    
    if len(m30) < 50 or len(h1) < 30:
        return f"[ERROR] 数据不足: M30={len(m30)}, H1={len(h1)}"
    
    bid = current_price.get('bid', m30[-1]['close'])
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # 2. Compute factor scores
    factor_data = compute_factors(m30, h1, current_price)
    if not factor_data:
        return "[ERROR] 因子计算失败"
    
    # 3. Log factor scores
    log_factor_scores(factor_data)
    f = factor_data['factors']
    
    # 4. Manage simulation positions
    state = load_state()
    state['_last_price'] = bid
    
    report = f"""
╔══════════════════════════════════════════════════╗
║  XAUUSD 手动策略模拟报告  {now_str}  ║
╚══════════════════════════════════════════════════╝

当前金价: {bid:.2f}
模拟资金: ${state['capital']:.2f} | 交易总数: {state['total_trades']}
胜率: {round(state['wins']/state['total_trades']*100, 1) if state['total_trades']>0 else 0}%
胜 {state['wins']} / 负 {state['losses']}

── 实时信号 ──"""

    for name, scores in f.items():
        report += f"\n  {name:<12} → 信号:{str(scores['信号']):>5}  评分:{scores['评分']}  {scores['说明']}"
    
    report += f"""

── 当前模拟持仓 ──"""
    if state['positions']:
        for p in state['positions']:
            dir_arrow = "▲" if p['direction'] == "BUY" else "▼"
            diff = (bid - p['open_price']) if p['direction'] == "BUY" else (p['open_price'] - bid)
            fl = round(diff * FIXED_LOT * 100, 2)
            report += f"\n  {dir_arrow} {p['strategy']} {p['direction']} @ {p['open_price']}  浮盈: ${fl:.2f}"
    else:
        report += "\n  无持仓"
    
    # 5. Check SL/TP and close conditions
    # Since we don't set SL/TP in this version, check for EXIT signals
    active_strats = {p['strategy'] for p in state['positions'] if p['strategy'] in f}
    
    for pos in list(state['positions']):
        strat_name = pos['strategy']
        if strat_name in f and f[strat_name]['信号'] == 'EXIT':
            pnl = close_position(state, pos, bid, f"信号出场: {f[strat_name]['说明']}", now_str)
            report += f"\n  📴 平仓 {strat_name} {pos['direction']} @ {bid:.2f} PnL=${pnl:.2f}"
    
    # 6. Check new entry signals
    for name, scores in f.items():
        if scores['信号'] in ('BUY', 'SELL') and name not in active_strats:
            # Check no opposite position exists
            has_opposite = any(p['strategy'] == name for p in state['positions'])
            if not has_opposite:
                open_position(state, name, scores['信号'], bid, now_str)
                arrow = "▲" if scores['信号'] == "BUY" else "▼"
                report += f"\n  🆕 {arrow} 开仓 {name} {scores['信号']} @ {bid:.2f}"
    
    # 7. Update equity
    floating = 0
    for p in state['positions']:
        diff = (bid - p['open_price']) if p['direction'] == "BUY" else (p['open_price'] - bid)
        floating += diff * FIXED_LOT * 100
    state['equity'] = round(state['capital'] + floating, 2)
    save_state(state)
    
    # 8. Check if anything changed (quiet mode support)
    had_actions = bool([l for l in report.split('\n') if '🆕' in l or '📴' in l or '🔴' in l or '🟢' in l])
    
    # 9. Generate Excel
    excel_path = generate_excel()
    
    # Quiet mode: suppress output when nothing happened (cron watchdog pattern)
    quiet = os.environ.get('SIM_QUIET', '0') == '1'
    if quiet and not had_actions:
        return ""  # Silent - nothing new
    
    report += f"""
  模拟权益: ${state['equity']:.2f} (浮动: ${floating:.2f})
  Excel报表: {excel_path}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"""
    
    return report


if __name__ == '__main__':
    result = main()
    print(result)
